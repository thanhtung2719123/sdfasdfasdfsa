# -*- coding: utf-8 -*-
"""
VNStock Premium Web Application Server
Powered by FastAPI
"""

import os
import sys
import time
import queue
import threading
import io
import sqlite3
try:
    import libsql
except ImportError:
    import sqlite3 as libsql
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime
from fastapi import FastAPI, BackgroundTasks, HTTPException, Query
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List, Dict, Any

# Ensure Vietnamese/UTF-8 print capability
sys.stdout.reconfigure(encoding='utf-8')

# Expose API Key Setup
try:
    from main import register_user
    register_user(api_key='vnstock_74f708b54d2a500d9fc23da9967a4cf5')
except Exception as e:
    print(f"Warning: could not register API key: {e}")

# Import existing codebase modules
from data_engine import DataEngine
from backtester import Backtester
from scanner import VN302, VN302_INDUSTRIES, SCAN_LIST, check_divergence

# Loại bỏ các cổ phiếu đã bị hủy niêm yết hoặc tạm ngừng giao dịch kéo dài (tránh treo API và làm bẩn log)
DELISTED_OR_SUSPENDED = {'ITA', 'TAR', 'LTG', 'POM', 'PTI', 'IBC', 'FLC', 'ROS', 'AMD', 'HAI', 'ART', 'DLG', 'TGG', 'BIX', 'SBT', 'VSN', 'DNW', 'PSH', 'TCD', 'CC1', 'DAG', 'VGG', 'PGS', 'SFG', 'TMS', 'SRC', 'STG', 'SZL', 'NVT'}
VN302 = [s for s in VN302 if s not in DELISTED_OR_SUSPENDED]
for ind, symbols in list(VN302_INDUSTRIES.items()):
    VN302_INDUSTRIES[ind] = [s for s in symbols if s not in DELISTED_OR_SUSPENDED]
SCAN_LIST = [s for s in SCAN_LIST if s not in DELISTED_OR_SUSPENDED]

def get_db_conn():
    db_url = os.getenv("TURSO_URL")
    db_token = os.getenv("TURSO_AUTH_TOKEN")
    
    if db_url and db_token:
        return libsql.connect(db_url, auth_token=db_token)
        
    conn = sqlite3.connect("market_cache.db", timeout=30.0)
    # Kích hoạt chế độ WAL (Write-Ahead Logging) giúp nhiều luồng đọc/ghi đồng thời không bị khóa DB
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
    except Exception as e:
        print(f"Warning: could not set PRAGMA: {e}")
    return conn

from vnstock import Company

CACHE_DIR = "data_cache"
if os.getenv("VERCEL") != "1":
    os.makedirs(CACHE_DIR, exist_ok=True)

# Add a set for dead symbols to prevent repeated API calls

# Initialize FastAPI app
app = FastAPI(title="VNStock TA Suite API", version="1.0.0")

# Global Application State
class AppState:
    def __init__(self):
        self.source = 'KBS'
        self.engine = DataEngine(source=self.source)
        
        # Threads locks and state containers
        self.scan_lock = threading.Lock()
        self.scan_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "results": [],
            "error": None
        }
        
        self.anti_lock = threading.Lock()
        self.anti_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "signals": [],
            "stats": {},
            "error": None
        }
        
        self.watch_lock = threading.Lock()
        self.watch_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "watchlist": [],
            "error": None
        }
        
        self.liq_lock = threading.Lock()
        self.liq_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "results": [],
            "summary": {},
            "error": None,
            "date": "2026-05-18"
        }
        
        self.cap_lock = threading.Lock()
        self.cap_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "results": [],
            "industries_summary": [],
            "top_10": [],
            "error": None,
            "date": "2026-05-18"
        }
        
        self.sync_lock = threading.Lock()
        self.sync_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "error": None
        }
        
        self.cap_range_lock = threading.Lock()
        self.cap_range_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 0,
            "results": [],
            "details_by_date": {},
            "error": None
        }
        
        self.shares_crawler_lock = threading.Lock()
        self.shares_crawler_state = {
            "running": False,
            "progress": 0.0,
            "current": 0,
            "total": 302,
            "completed": False
        }
state = AppState()

# Pydantic schema for Backtest request
class BacktestParams(BaseModel):
    symbol: str
    start_date: str
    end_date: str
    initial_capital: float
    position_size: float = 100.0
    rsi_period: int = 14
    buy_threshold: float = 30.0
    sell_threshold: float = 70.0
    stop_loss: Optional[str] = "none"
    take_profit: Optional[str] = "none"

# Serve Frontend static index page
@app.get("/")
def read_index():
    index_path = os.path.join("static", "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return HTMLResponse("<h1>static/index.html not found.</h1>")

# --- SOURCE MANAGEMENT ---
@app.post("/api/source")
def set_data_source(source: str):
    if source.upper() not in ['KBS', 'MSN', 'VCI']:
        raise HTTPException(status_code=400, detail="Unsupported source. Use KBS, MSN, or VCI.")
    state.source = source.upper()
    state.engine = DataEngine(source=state.source)
    print(f"DataEngine source updated to: {state.source}")
    return {"status": "success", "source": state.source}

# --- TICKERS LIST ---
@app.get("/api/tickers")
def get_tickers():
    return {"tickers": VN302}

# --- 1. RSI BACKTESTER ROUTE ---
@app.post("/api/backtest")
def run_rsi_backtest(params: BacktestParams):
    symbol = params.symbol.upper()
    if symbol not in VN302:
        raise HTTPException(status_code=400, detail=f"Symbol {symbol} not in VN302 universe.")
        
    try:
        # Parse SL & TP options
        sl_val = params.stop_loss.strip().lower()
        stop_loss = float(sl_val) if sl_val != 'none' and sl_val != '' else None
        
        tp_val = params.take_profit.strip().lower()
        take_profit = float(tp_val) if tp_val != 'none' and tp_val != '' else None
        
        # Instantiate Backtester
        bt = Backtester(
            symbol=symbol,
            start_date=params.start_date,
            end_date=params.end_date,
            initial_capital=params.initial_capital,
            rsi_period=params.rsi_period,
            buy_threshold=params.buy_threshold,
            sell_threshold=params.sell_threshold,
            stop_loss=stop_loss,
            take_profit=take_profit,
            position_size=params.position_size,
            engine=state.engine
        )
        
        trades, metrics, df_curves = bt.run_backtest()
        
        # Format trades for UI
        ui_trades = []
        for t in trades:
            ui_trades.append({
                "entry_date": t['entry_date'].strftime('%Y-%m-%d'),
                "exit_date": t['exit_date'].strftime('%Y-%m-%d'),
                "entry_price": float(t['entry_price']),
                "exit_price": float(t['exit_price']),
                "profit": float(t['profit']),
                "profit_pct": float(t['profit_pct'])
            })
            
        # Format equity curve for Chart.js
        ui_equity = []
        if df_curves is not None and not df_curves.empty:
            for _, r in df_curves.iterrows():
                ui_equity.append({
                    "date": r['time'].strftime('%Y-%m-%d'),
                    "value": float(r['equity'])
                })
                
        return {
            "symbol": symbol,
            "metrics": metrics,
            "trades": ui_trades,
            "equity_curve": ui_equity
        }
        
    except Exception as e:
        print(f"Backtest error: {e}")
        return {"error": str(e)}

# --- 2. MULTI-THREADED SCANNER BACKGROUND TASK ---
def run_scan_worker():
    with state.scan_lock:
        state.scan_state["running"] = True
        state.scan_state["progress"] = 0.0
        state.scan_state["current"] = 0
        state.scan_state["total"] = len(SCAN_LIST)
        state.scan_state["results"] = []
        state.scan_state["error"] = None
        
    total = len(SCAN_LIST)
    fetch_start_date = '2025-01-01'
    return_ref_date = '2026-03-23'
    
    q = queue.Queue()
    for symbol in SCAN_LIST:
        q.put(symbol)
        
    completed = 0
    results_list = []
    results_lock = threading.Lock()
    progress_lock = threading.Lock()
    
    def worker():
        nonlocal completed
        while state.scan_state["running"]:
            try:
                symbol = q.get_nowait()
            except queue.Empty:
                break
                
            try:
                df_long = state.engine.get_history(symbol, start=fetch_start_date)
                
                if df_long.empty or len(df_long) < 10:
                    q.task_done()
                    with progress_lock:
                        completed += 1
                        state.scan_state["current"] = completed
                        state.scan_state["progress"] = (completed / total) * 100
                    continue

                close = df_long['close']
                current_price = close.iloc[-1]
                
                # Performance since March 23, 2026
                df_since_ref = df_long[df_long['time'] >= return_ref_date]
                if not df_since_ref.empty:
                    price_start = df_since_ref['close'].iloc[0]
                    return_since_ref = (current_price / price_start - 1) * 100
                else:
                    return_since_ref = 0
                
                ma20 = close.rolling(20).mean().iloc[-1] if len(close) >= 20 else 0
                ma50 = close.rolling(50).mean().iloc[-1] if len(close) >= 50 else 0
                ma100 = close.rolling(100).mean().iloc[-1] if len(close) >= 100 else 0
                ma200 = close.rolling(200).mean().iloc[-1] if len(close) >= 200 else 0
                
                window_52w = 252
                if len(close) >= window_52w:
                    rolling_52w = close.iloc[-window_52w:]
                    high_52w = rolling_52w.max()
                    low_52w = rolling_52w.min()
                    high_idx = rolling_52w.idxmax()
                    breakout_date = df_long.loc[high_idx, 'time'].strftime('%Y-%m-%d')
                else:
                    high_52w = close.max()
                    low_52w = close.min()
                    high_idx = close.idxmax()
                    breakout_date = df_long.loc[high_idx, 'time'].strftime('%Y-%m-%d')

                delta = close.diff()
                gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
                loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
                rs = gain / loss
                rsi = 100 - (100 / (1 + rs))
                current_rsi = rsi.iloc[-1] if pd.notna(rsi.iloc[-1]) else 50
                
                ema12 = close.ewm(span=12, adjust=False).mean()
                ema26 = close.ewm(span=26, adjust=False).mean()
                macd = ema12 - ema26
                macd_signal = macd.ewm(span=9, adjust=False).mean()
                macd_hist = macd - macd_signal
                
                rsi_div = check_divergence(close, rsi)
                macd_div = check_divergence(close, macd_hist)
                
                if len(macd_hist) > 1:
                    macd_status = "Cross Up" if macd_hist.iloc[-1] > 0 and macd_hist.iloc[-2] <= 0 else \
                                  "Cross Down" if macd_hist.iloc[-1] < 0 and macd_hist.iloc[-2] >= 0 else \
                                  "Positive" if macd_hist.iloc[-1] > 0 else "Negative"
                else:
                    macd_status = "N/A"

                industry = "Index"
                if symbol != 'VNINDEX':
                    for ind, symbols in VN302_INDUSTRIES.items():
                        if symbol in symbols:
                            industry = ind
                            break

                # 52w Breakout classifications
                breakout_status = "Normal"
                if current_price >= high_52w:
                    breakout_status = "High Breakout"
                elif current_price <= low_52w:
                    breakout_status = "Low Breakout"

                row_data = {
                    "Ticker": symbol,
                    "Industry": industry,
                    "Price": float(current_price),
                    "Return_2026_03_23": float(return_since_ref),
                    "MA20": 1 if current_price >= ma20 else 0,
                    "MA50": 1 if current_price >= ma50 else 0,
                    "MA100": 1 if current_price >= ma100 else 0,
                    "MA200": 1 if current_price >= ma200 else 0,
                    "High_52w": float(high_52w),
                    "Low_52w": float(low_52w),
                    "Breakout_Date": breakout_date,
                    "Breakout": breakout_status,
                    "RSI": float(current_rsi),
                    "RSI_Divergence": rsi_div,
                    "MACD": macd_status,
                    "MACD_Divergence": macd_div
                }
                
                with results_lock:
                    results_list.append(row_data)
                    
            except Exception as e:
                print(f"Error scanning symbol {symbol}: {e}")
                
            q.task_done()
            with progress_lock:
                completed += 1
                state.scan_state["current"] = completed
                state.scan_state["progress"] = (completed / total) * 100

    num_workers = 8
    threads = []
    for _ in range(num_workers):
        t = threading.Thread(target=worker)
        t.start()
        threads.append(t)
        
    for t in threads:
        t.join()
        
    with state.scan_lock:
        state.scan_state["results"] = results_list
        state.scan_state["progress"] = 100.0
        state.scan_state["running"] = False

# REST routes for Scanner
@app.post("/api/scan/start")
def start_market_scan(background_tasks: BackgroundTasks):
    if state.scan_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_scan_worker)
    return {"status": "started"}

@app.post("/api/scan/stop")
def stop_market_scan():
    state.scan_state["running"] = False
    return {"status": "stopped"}

@app.get("/api/scan/status")
def get_scan_status():
    return {
        "running": state.scan_state["running"],
        "progress": state.scan_state["progress"],
        "current": state.scan_state["current"],
        "total": state.scan_state["total"]
    }

@app.get("/api/scan/results")
def get_scan_results():
    return {"results": state.scan_state["results"]}

# --- 3. SECTOR HEATMAP API ---
@app.get("/api/heatmap")
def get_sector_heatmap():
    results = state.scan_state["results"]
    if not results:
        return {"sectors": {}}
        
    df = pd.DataFrame(results)
    # Filter out Index
    df = df[df['Ticker'] != 'VNINDEX']
    
    sectors_summary = {}
    for ind, group in df.groupby('Industry'):
        above_ma50 = int((group['MA50'] == 1).sum())
        total_count = int(len(group))
        
        sectors_summary[ind] = {
            "avg_return": float(group['Return_2026_03_23'].mean()),
            "above_ma50_pct": float(above_ma50 / total_count if total_count > 0 else 0),
            "above_ma50_count": above_ma50,
            "total_count": total_count
        }
        
    return {"sectors": sectors_summary}


def calculate_percentile(values, current_value):
    clean_values = [float(v) for v in values if pd.notna(v)]
    if not clean_values or current_value is None or pd.isna(current_value):
        return None
    return round((sum(v <= current_value for v in clean_values) / len(clean_values)) * 100.0, 1)


SECTOR_FLOW_PERIODS = {
    "1d": {"sessions": 1, "label": "1 ngÃ y"},
    "3d": {"sessions": 3, "label": "3 ngÃ y"},
    "5d": {"sessions": 5, "label": "5 ngÃ y"},
    "2w": {"sessions": 10, "label": "2 tuáº§n"},
    "1m": {"sessions": 21, "label": "1 thÃ¡ng"},
    "1q": {"sessions": 63, "label": "1 quÃ½"},
    "1y": {"sessions": 252, "label": "1 nÄƒm"},
    "60d": {"sessions": 60, "label": "60 phiÃªn"},
}


def build_sector_flow_snapshot(years: int = 3, period: str = "60d"):
    period_config = SECTOR_FLOW_PERIODS.get(period, SECTOR_FLOW_PERIODS["60d"])
    lookback_sessions = period_config["sessions"]
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(time) FROM historical_prices")
    max_date_str = cursor.fetchone()[0]
    if not max_date_str:
        conn.close()
        return {"error": "Database chưa có dữ liệu giá. Bấm Đồng bộ Data 3 năm trước."}

    max_dt = datetime.strptime(max_date_str, "%Y-%m-%d")
    start_date = (max_dt - pd.DateOffset(years=years)).strftime("%Y-%m-%d")

    cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
    shares_map = {row[0]: row[1] for row in cursor.fetchall()}
    cursor.execute("""
        SELECT symbol, time, close, volume
        FROM historical_prices
        WHERE time >= ? AND time <= ?
    """, (start_date, max_date_str))
    rows = cursor.fetchall()
    conn.close()

    records = []
    active_symbols = set(VN302)
    for symbol, dt, close_val, volume in rows:
        if symbol not in active_symbols:
            continue
        shares = get_outstanding_shares(symbol, shares_map)
        if not shares or close_val is None or close_val <= 0 or volume is None:
            continue
        price_vnd = normalize_price(close_val)
        industry = get_symbol_industry(symbol)
        records.append({
            "Date": dt,
            "Industry": industry,
            "Ticker": symbol,
            "GTGDBillion": (price_vnd * int(volume)) / 1_000_000_000,
            "CapBillion": (price_vnd * shares) / 1_000_000_000
        })

    if not records:
        return {"error": "Không đủ dữ liệu ngành trong database."}

    df = pd.DataFrame(records)
    grouped = (
        df.groupby(["Date", "Industry"], as_index=False)
        .agg(
            GTGDBillion=("GTGDBillion", "sum"),
            CapBillion=("CapBillion", "sum"),
            TickerCount=("Ticker", "nunique")
        )
    )
    market_gtgd = grouped.groupby("Date")["GTGDBillion"].sum().rename("MarketGTGDBillion")
    grouped = grouped.merge(market_gtgd, on="Date", how="left")
    grouped["GTGDCapPct"] = np.where(grouped["CapBillion"] > 0, grouped["GTGDBillion"] / grouped["CapBillion"] * 100.0, np.nan)
    grouped["MarketSharePct"] = np.where(grouped["MarketGTGDBillion"] > 0, grouped["GTGDBillion"] / grouped["MarketGTGDBillion"] * 100.0, np.nan)
    grouped = grouped.sort_values(["Industry", "Date"]).reset_index(drop=True)
    grouped["GTGD_MA20"] = grouped.groupby("Industry")["GTGDBillion"].transform(lambda s: s.rolling(20, min_periods=1).mean())
    grouped["GTGD_MA60"] = grouped.groupby("Industry")["GTGDBillion"].transform(lambda s: s.rolling(60, min_periods=1).mean())
    grouped["Ratio_MA20"] = grouped.groupby("Industry")["GTGDCapPct"].transform(lambda s: s.rolling(20, min_periods=1).mean())
    grouped["Ratio_MA60"] = grouped.groupby("Industry")["GTGDCapPct"].transform(lambda s: s.rolling(60, min_periods=1).mean())
    grouped["Share_MA20"] = grouped.groupby("Industry")["MarketSharePct"].transform(lambda s: s.rolling(20, min_periods=1).mean())
    grouped["Share_MA60"] = grouped.groupby("Industry")["MarketSharePct"].transform(lambda s: s.rolling(60, min_periods=1).mean())

    all_dates = sorted(grouped["Date"].unique().tolist())
    latest_date = grouped["Date"].max()
    latest_idx = all_dates.index(latest_date)
    compare_idx = max(0, latest_idx - lookback_sessions)
    compare_date = all_dates[compare_idx]
    latest = grouped[grouped["Date"] == latest_date].copy()
    compare = grouped[grouped["Date"] == compare_date].copy()
    compare_by_industry = compare.set_index("Industry").to_dict("index")

    ticker_latest = df[df["Date"] == latest_date].copy()
    ticker_compare = df[df["Date"] == compare_date].copy()
    ticker_compare_gtgd = ticker_compare.set_index("Ticker")["GTGDBillion"].to_dict()
    latest_industry_gtgd = latest.set_index("Industry")["GTGDBillion"].to_dict()

    highlight_tickers = {}
    for industry, industry_group in ticker_latest.groupby("Industry"):
        ticker_rows = []
        for _, ticker_row in industry_group.iterrows():
            current_gtgd = float(ticker_row["GTGDBillion"])
            base_gtgd = float(ticker_compare_gtgd.get(ticker_row["Ticker"], 0) or 0)
            change_gtgd = current_gtgd - base_gtgd
            industry_gtgd = float(latest_industry_gtgd.get(industry, 0) or 0)
            ticker_rows.append({
                "Ticker": ticker_row["Ticker"],
                "GTGDBillion": round(current_gtgd, 2),
                "ChangeBillion": round(change_gtgd, 2),
                "ChangePct": round(((current_gtgd / base_gtgd) - 1) * 100.0, 2) if base_gtgd > 0 else None,
                "IndustrySharePct": round((current_gtgd / industry_gtgd) * 100.0, 2) if industry_gtgd > 0 else None
            })
        positive_rows = [item for item in ticker_rows if item["ChangeBillion"] > 0]
        ranked_tickers = positive_rows if positive_rows else ticker_rows
        highlight_tickers[industry] = sorted(
            ranked_tickers,
            key=lambda item: (item["ChangeBillion"], item["GTGDBillion"]),
            reverse=True
        )[:3]

    output_rows = []
    for _, row in latest.iterrows():
        hist = grouped[grouped["Industry"] == row["Industry"]]
        compare_row = compare_by_industry.get(row["Industry"], {})
        compare_gtgd = compare_row.get("GTGDBillion")
        compare_cap = compare_row.get("CapBillion")
        compare_ratio = compare_row.get("GTGDCapPct")
        compare_share = compare_row.get("MarketSharePct")
        ratio_ma20 = float(row["Ratio_MA20"]) if pd.notna(row["Ratio_MA20"]) else None
        ratio_ma60 = float(row["Ratio_MA60"]) if pd.notna(row["Ratio_MA60"]) else None
        share_ma20 = float(row["Share_MA20"]) if pd.notna(row["Share_MA20"]) else None
        share_ma60 = float(row["Share_MA60"]) if pd.notna(row["Share_MA60"]) else None
        output_rows.append({
            "Industry": row["Industry"],
            "TickerCount": int(row["TickerCount"]),
            "GTGDBillion": round(float(row["GTGDBillion"]), 2),
            "CapBillion": round(float(row["CapBillion"]), 2),
            "GTGDCapPct": round(float(row["GTGDCapPct"]), 4),
            "MarketSharePct": round(float(row["MarketSharePct"]), 4),
            "GTGDPercentile": calculate_percentile(hist["GTGDBillion"], row["GTGDBillion"]),
            "CapPercentile": calculate_percentile(hist["CapBillion"], row["CapBillion"]),
            "GTGDCapPercentile": calculate_percentile(hist["GTGDCapPct"], row["GTGDCapPct"]),
            "MarketSharePercentile": calculate_percentile(hist["MarketSharePct"], row["MarketSharePct"]),
            "GTGDMA20": round(float(row["GTGD_MA20"]), 2),
            "GTGDMA60": round(float(row["GTGD_MA60"]), 2),
            "RatioMA20": round(ratio_ma20, 4) if ratio_ma20 is not None else None,
            "RatioMA60": round(ratio_ma60, 4) if ratio_ma60 is not None else None,
            "ShareMA20": round(share_ma20, 4) if share_ma20 is not None else None,
            "ShareMA60": round(share_ma60, 4) if share_ma60 is not None else None,
            "RatioVsMA20Pct": round(((row["GTGDCapPct"] / ratio_ma20) - 1) * 100.0, 2) if ratio_ma20 and ratio_ma20 > 0 else None,
            "RatioVsMA60Pct": round(((row["GTGDCapPct"] / ratio_ma60) - 1) * 100.0, 2) if ratio_ma60 and ratio_ma60 > 0 else None,
            "ShareVsMA20Pct": round(row["MarketSharePct"] - share_ma20, 4) if share_ma20 is not None else None,
            "ShareVsMA60Pct": round(row["MarketSharePct"] - share_ma60, 4) if share_ma60 is not None else None,
            "PeriodGTGDChangeBillion": round(float(row["GTGDBillion"]) - float(compare_gtgd), 2) if compare_gtgd is not None and pd.notna(compare_gtgd) else None,
            "PeriodGTGDChangePct": round(((float(row["GTGDBillion"]) / float(compare_gtgd)) - 1) * 100.0, 2) if compare_gtgd is not None and pd.notna(compare_gtgd) and float(compare_gtgd) > 0 else None,
            "PeriodCapChangePct": round(((float(row["CapBillion"]) / float(compare_cap)) - 1) * 100.0, 2) if compare_cap is not None and pd.notna(compare_cap) and float(compare_cap) > 0 else None,
            "PeriodRatioChangePct": round(((float(row["GTGDCapPct"]) / float(compare_ratio)) - 1) * 100.0, 2) if compare_ratio is not None and pd.notna(compare_ratio) and float(compare_ratio) > 0 else None,
            "PeriodShareChangePct": round(float(row["MarketSharePct"]) - float(compare_share), 4) if compare_share is not None and pd.notna(compare_share) else None,
            "CompareMarketSharePct": round(float(compare_share), 4) if compare_share is not None and pd.notna(compare_share) else None,
            "HighlightTickers": highlight_tickers.get(row["Industry"], [])
        })

    output_rows = sorted(output_rows, key=lambda item: item["GTGDBillion"], reverse=True)
    for rank, item in enumerate(output_rows, 1):
        item["Rank"] = rank

    daily_leaders = []
    sorted_dates = sorted(grouped["Date"].unique().tolist())
    for idx in range(1, len(sorted_dates)):
        current_date = sorted_dates[idx]
        previous_date = sorted_dates[idx - 1]
        current_tickers = df[df["Date"] == current_date]
        previous_tickers = df[df["Date"] == previous_date].set_index("Ticker")["GTGDBillion"].to_dict()
        ticker_changes = []
        for _, ticker_row in current_tickers.iterrows():
            prev_gtgd = float(previous_tickers.get(ticker_row["Ticker"], 0) or 0)
            current_gtgd = float(ticker_row["GTGDBillion"])
            ticker_changes.append({
                "Ticker": ticker_row["Ticker"],
                "Industry": ticker_row["Industry"],
                "GTGDBillion": round(current_gtgd, 2),
                "ChangeBillion": round(current_gtgd - prev_gtgd, 2),
                "ChangePct": round(((current_gtgd / prev_gtgd) - 1) * 100.0, 2) if prev_gtgd > 0 else None
            })

        current_sectors = grouped[grouped["Date"] == current_date].set_index("Industry")
        previous_sectors = grouped[grouped["Date"] == previous_date].set_index("Industry")
        sector_changes = []
        for industry, sector_row in current_sectors.iterrows():
            prev_share = previous_sectors.loc[industry]["MarketSharePct"] if industry in previous_sectors.index else 0
            sector_changes.append({
                "Industry": industry,
                "ShareChangePct": round(float(sector_row["MarketSharePct"]) - float(prev_share), 4),
                "MarketSharePct": round(float(sector_row["MarketSharePct"]), 4)
            })

        top_in_ticker = max(ticker_changes, key=lambda item: item["ChangeBillion"]) if ticker_changes else None
        top_out_ticker = min(ticker_changes, key=lambda item: item["ChangeBillion"]) if ticker_changes else None
        top_in_sector = max(sector_changes, key=lambda item: item["ShareChangePct"]) if sector_changes else None
        top_out_sector = min(sector_changes, key=lambda item: item["ShareChangePct"]) if sector_changes else None
        daily_leaders.append({
            "Date": current_date,
            "PrevDate": previous_date,
            "TopInTicker": top_in_ticker,
            "TopOutTicker": top_out_ticker,
            "TopInSector": top_in_sector,
            "TopOutSector": top_out_sector
        })

    return {
        "start_date": start_date,
        "end_date": latest_date,
        "compare_date": compare_date,
        "period": period,
        "period_label": period_config["label"],
        "period_sessions": lookback_sessions,
        "years": years,
        "rows": output_rows,
        "daily_leaders": list(reversed(daily_leaders[-60:])),
        "history_points": int(grouped["Date"].nunique()),
        "source": "database"
    }


def build_sector_flow_history(years: int = 3):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(time) FROM historical_prices")
    max_date_str = cursor.fetchone()[0]
    if not max_date_str:
        conn.close()
        return {"error": "Database chÆ°a cÃ³ dá»¯ liá»‡u giÃ¡."}

    max_dt = datetime.strptime(max_date_str, "%Y-%m-%d")
    start_date = (max_dt - pd.DateOffset(years=years)).strftime("%Y-%m-%d")
    cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
    shares_map = {row[0]: row[1] for row in cursor.fetchall()}
    cursor.execute("""
        SELECT symbol, time, close, volume
        FROM historical_prices
        WHERE time >= ? AND time <= ?
    """, (start_date, max_date_str))
    rows = cursor.fetchall()
    conn.close()

    records = []
    for symbol, dt, close_val, volume in rows:
        if symbol not in VN302:
            continue
        shares = get_outstanding_shares(symbol, shares_map)
        if not shares or close_val is None or close_val <= 0 or volume is None:
            continue
        price_vnd = normalize_price(close_val)
        industry = get_symbol_industry(symbol)
        records.append({
            "Date": dt,
            "Industry": industry,
            "Ticker": symbol,
            "GTGDBillion": (price_vnd * int(volume)) / 1_000_000_000,
            "CapBillion": (price_vnd * shares) / 1_000_000_000
        })

    if not records:
        return {"error": "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u ngÃ nh trong database."}

    df = pd.DataFrame(records)
    grouped = (
        df.groupby(["Date", "Industry"], as_index=False)
        .agg(
            GTGDBillion=("GTGDBillion", "sum"),
            CapBillion=("CapBillion", "sum"),
            TickerCount=("Ticker", "nunique")
        )
    )
    market_gtgd = grouped.groupby("Date")["GTGDBillion"].sum().rename("MarketGTGDBillion")
    grouped = grouped.merge(market_gtgd, on="Date", how="left")
    grouped["GTGDCapPct"] = np.where(grouped["CapBillion"] > 0, grouped["GTGDBillion"] / grouped["CapBillion"] * 100.0, np.nan)
    grouped["MarketSharePct"] = np.where(grouped["MarketGTGDBillion"] > 0, grouped["GTGDBillion"] / grouped["MarketGTGDBillion"] * 100.0, np.nan)
    grouped = grouped.sort_values(["Industry", "Date"]).reset_index(drop=True)

    industries_from_map = [ind for ind in VN302_INDUSTRIES.keys() if ind in set(grouped["Industry"])]
    extra_industries = sorted(set(grouped["Industry"]) - set(industries_from_map))
    industries = industries_from_map + extra_industries
    dates_desc = sorted(grouped["Date"].unique().tolist(), reverse=True)

    rows_out = []
    for dt in dates_desc:
        day = grouped[grouped["Date"] == dt].set_index("Industry")
        values = {}
        for industry in industries:
            if industry in day.index:
                item = day.loc[industry]
                values[industry] = {
                    "GTGDBillion": round(float(item["GTGDBillion"]), 2),
                    "CapBillion": round(float(item["CapBillion"]), 2),
                    "GTGDCapPct": round(float(item["GTGDCapPct"]), 4),
                    "MarketSharePct": round(float(item["MarketSharePct"]), 4),
                }
            else:
                values[industry] = {}
        rows_out.append({"Date": dt, "values": values})

    def metric_stats(metric_col):
        stat_rows = {
            "Average 20 phiÃªn": {},
            "Average 60 phiÃªn": {},
            "Average 250 phiÃªn": {},
            "Min 52 tuáº§n": {},
            "Max 52 tuáº§n": {},
            "Average 1 tuáº§n": {},
            "Average 52 tuáº§n": {},
            "5%": {},
            "50%": {},
            "95%": {},
        }
        for industry in industries:
            s = grouped[grouped["Industry"] == industry].sort_values("Date")[metric_col].dropna()
            if s.empty:
                continue
            stat_rows["Average 20 phiÃªn"][industry] = round(float(s.tail(20).mean()), 4)
            stat_rows["Average 60 phiÃªn"][industry] = round(float(s.tail(60).mean()), 4)
            stat_rows["Average 250 phiÃªn"][industry] = round(float(s.tail(250).mean()), 4)
            last_252 = s.tail(252)
            stat_rows["Min 52 tuáº§n"][industry] = round(float(last_252.min()), 4)
            stat_rows["Max 52 tuáº§n"][industry] = round(float(last_252.max()), 4)
            stat_rows["Average 1 tuáº§n"][industry] = round(float(s.tail(5).mean()), 4)
            stat_rows["Average 52 tuáº§n"][industry] = round(float(last_252.mean()), 4)
            stat_rows["5%"][industry] = round(float(s.quantile(0.05)), 4)
            stat_rows["50%"][industry] = round(float(s.quantile(0.50)), 4)
            stat_rows["95%"][industry] = round(float(s.quantile(0.95)), 4)
        return stat_rows

    return {
        "start_date": start_date,
        "end_date": max_date_str,
        "years": years,
        "industries": industries,
        "rows": rows_out,
        "stats": {
            "gtgd": metric_stats("GTGDBillion"),
            "cap": metric_stats("CapBillion"),
            "ratio": metric_stats("GTGDCapPct"),
            "share": metric_stats("MarketSharePct"),
        },
        "source": "database"
    }


@app.get("/api/sector-flow/results")
def get_sector_flow_results(years: int = Query(3, ge=1, le=5), period: str = Query("60d")):
    return build_sector_flow_snapshot(years=years, period=period)


@app.get("/api/sector-flow/history")
def get_sector_flow_history(years: int = Query(3, ge=1, le=5)):
    return build_sector_flow_history(years=years)


@app.get("/api/export/sector-flow")
def export_sector_flow_excel(years: int = Query(3, ge=1, le=5)):
    data = build_sector_flow_snapshot(years=years)
    if data.get("error"):
        raise HTTPException(status_code=400, detail=data["error"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sector Flow"
    headers = [
        "Rank", "Industry", "Tickers", "GTGD ngành (tỷ)", "Vốn hóa ngành (tỷ)",
        "GTGD/Vốn hóa (%)", "Tỷ trọng GTGD (%)", "Pctl GTGD", "Pctl Vốn hóa",
        "Pctl GTGD/VH", "Pctl Tỷ trọng", "GTGD MA20", "GTGD MA60",
        "GTGD/VH MA20", "GTGD/VH MA60", "So với MA20 (%)", "So với MA60 (%)"
    ]
    ws.append(headers)
    for item in data["rows"]:
        ws.append([
            item["Rank"], item["Industry"], item["TickerCount"], item["GTGDBillion"],
            item["CapBillion"], item["GTGDCapPct"], item["MarketSharePct"],
            item["GTGDPercentile"], item["CapPercentile"], item["GTGDCapPercentile"],
            item["MarketSharePercentile"], item["GTGDMA20"], item["GTGDMA60"],
            item["RatioMA20"], item["RatioMA60"], item["RatioVsMA20Pct"],
            item["RatioVsMA60Pct"]
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sector_flow_{data['end_date']}_{years}y.xlsx"}
    )


@app.get("/api/export/sector-flow-matrix")
def export_sector_flow_matrix_excel(years: int = Query(3, ge=1, le=5)):
    data = build_sector_flow_history(years=years)
    if data.get("error"):
        raise HTTPException(status_code=400, detail=data["error"])

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    metric_defs = {
        "gtgd": {
            "sheet": "GTGD nganh",
            "title": "Giá trị giao dịch (tỷ đồng)",
            "key": "GTGDBillion",
            "summary": ["Average 20 phiÃªn", "Average 60 phiÃªn", "Average 250 phiÃªn"],
            "number_format": "#,##0",
            "percent": False,
            "heat": False
        },
        "cap": {
            "sheet": "Von hoa nganh",
            "title": "Vốn hóa ngành (tỷ đồng)",
            "key": "CapBillion",
            "summary": [],
            "number_format": "#,##0",
            "percent": False,
            "heat": False
        },
        "ratio": {
            "sheet": "GTGD von hoa",
            "title": "GTGD / Vốn hóa ngành",
            "key": "GTGDCapPct",
            "summary": [],
            "number_format": "0.00%",
            "percent": True,
            "heat": True
        },
        "share": {
            "sheet": "Ty trong nganh",
            "title": "Giá trị giao dịch theo ngày (khớp lệnh)",
            "key": "MarketSharePct",
            "summary": ["Min 52 tuáº§n", "Max 52 tuáº§n", "Average 1 tuáº§n", "Average 52 tuáº§n", "Average 60 phiÃªn", "Average 20 phiÃªn", "5%", "50%", "95%"],
            "number_format": "0.00%",
            "percent": True,
            "heat": True
        }
    }

    def display_value(value, is_percent):
        if value is None or pd.isna(value):
            return None
        return float(value) / 100.0 if is_percent else round(float(value), 0)

    def add_sheet(metric_name, config):
        ws = wb.create_sheet(title=config["sheet"])
        row_idx = 1
        for label in config["summary"]:
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="FF0000" if metric_name == "share" else "000000")
            for col_idx, industry in enumerate(data["industries"], 2):
                raw_val = data["stats"][metric_name].get(label, {}).get(industry)
                cell = ws.cell(row=row_idx, column=col_idx, value=display_value(raw_val, config["percent"]))
                cell.number_format = config["number_format"]
                cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        if config["summary"]:
            row_idx += 1

        ws.cell(row=row_idx, column=1, value=config["title"])
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="FF0000")
        ws.cell(row=row_idx, column=1).fill = title_fill
        row_idx += 1

        header_row = row_idx
        ws.cell(row=header_row, column=1, value="Dates")
        for col_idx, industry in enumerate(data["industries"], 2):
            ws.cell(row=header_row, column=col_idx, value=industry)
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        data_start_row = header_row + 1
        for out_row_idx, day in enumerate(data["rows"], data_start_row):
            dt = datetime.strptime(day["Date"], "%Y-%m-%d")
            ws.cell(row=out_row_idx, column=1, value=f"{dt.day}/{dt.month}/{dt.year}")
            ws.cell(row=out_row_idx, column=1).font = Font(bold=True)
            ws.cell(row=out_row_idx, column=1).border = border
            for col_idx, industry in enumerate(data["industries"], 2):
                raw_val = day["values"].get(industry, {}).get(config["key"])
                cell = ws.cell(row=out_row_idx, column=col_idx, value=display_value(raw_val, config["percent"]))
                cell.number_format = config["number_format"]
                cell.alignment = Alignment(horizontal="right")
                cell.border = border

        if config["heat"]:
            for col_idx in range(2, len(data["industries"]) + 2):
                values = [
                    ws.cell(row=r, column=col_idx).value
                    for r in range(data_start_row, data_start_row + len(data["rows"]))
                    if ws.cell(row=r, column=col_idx).value is not None
                ]
                if not values:
                    continue
                low = np.quantile(values, 0.25)
                high = np.quantile(values, 0.75)
                for r in range(data_start_row, data_start_row + len(data["rows"])):
                    value = ws.cell(row=r, column=col_idx).value
                    if value is None:
                        continue
                    if value >= high:
                        ws.cell(row=r, column=col_idx).fill = green_fill
                    elif value <= low:
                        ws.cell(row=r, column=col_idx).fill = red_fill

        ws.freeze_panes = ws.cell(row=data_start_row, column=2).coordinate
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(data['industries']) + 1)}{data_start_row + len(data['rows']) - 1}"
        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, len(data["industries"]) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for metric_name, config in metric_defs.items():
        add_sheet(metric_name, config)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sector_flow_matrix_{data['end_date']}_{years}y.xlsx"}
    )


@app.get("/api/export/all-market-data")
def export_all_market_data_excel(years: int = Query(3, ge=1, le=5)):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(time) FROM historical_prices")
    max_date_str = cursor.fetchone()[0]
    if not max_date_str:
        conn.close()
        raise HTTPException(status_code=400, detail="Database chưa có dữ liệu giá.")

    max_dt = datetime.strptime(max_date_str, "%Y-%m-%d")
    start_date = (max_dt - pd.DateOffset(years=years)).strftime("%Y-%m-%d")
    cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
    shares_map = {row[0]: row[1] for row in cursor.fetchall()}
    cursor.execute("""
        SELECT symbol, time, close, volume
        FROM historical_prices
        WHERE time >= ? AND time <= ?
    """, (start_date, max_date_str))
    rows = cursor.fetchall()
    conn.close()

    ordered_tickers = []
    ticker_to_industry = {}
    for industry, symbols in VN302_INDUSTRIES.items():
        for symbol in symbols:
            if symbol in VN302:
                ordered_tickers.append(symbol)
                ticker_to_industry[symbol] = industry

    dates = sorted({row[1] for row in rows}, reverse=True)
    price_matrix = {}
    volume_matrix = {}
    for symbol, dt, close_val, volume in rows:
        if symbol not in VN302:
            continue
        price_matrix.setdefault(dt, {})[symbol] = close_val
        volume_matrix.setdefault(dt, {})[symbol] = volume

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

    def add_matrix_sheet(title, metric_type):
        ws = wb.create_sheet(title=title)
        ws.cell(row=1, column=1, value="Date")
        ws.cell(row=2, column=1, value="Industry")
        ws.cell(row=3, column=1, value="Ticker")
        for col_idx, symbol in enumerate(ordered_tickers, 2):
            ws.cell(row=1, column=col_idx, value=ticker_to_industry.get(symbol, "Khác"))
            ws.cell(row=2, column=col_idx, value=symbol)
            ws.cell(row=3, column=col_idx, value=metric_type)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, dt in enumerate(dates, 4):
            ws.cell(row=row_idx, column=1, value=dt)
            day_prices = price_matrix.get(dt, {})
            day_volumes = volume_matrix.get(dt, {})
            for col_idx, symbol in enumerate(ordered_tickers, 2):
                close_val = day_prices.get(symbol)
                volume = day_volumes.get(symbol)
                shares = get_outstanding_shares(symbol, shares_map)
                value = None
                if close_val is not None and close_val > 0:
                    price_vnd = normalize_price(close_val)
                    if metric_type == "LIQUIDITY_BN" and volume is not None:
                        value = (price_vnd * int(volume)) / 1_000_000_000
                    elif metric_type == "MARKET_CAP_BN" and shares:
                        value = (price_vnd * shares) / 1_000_000_000
                    elif metric_type == "GTGD_CAP_RATIO" and shares and volume is not None:
                        value = int(volume) / shares
                cell = ws.cell(row=row_idx, column=col_idx, value=round(value, 6) if value is not None else None)
                if metric_type == "GTGD_CAP_RATIO":
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "B4"
        ws.column_dimensions[get_column_letter(1)].width = 14
        for col_idx in range(2, len(ordered_tickers) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13

    add_matrix_sheet("Thanh khoan", "LIQUIDITY_BN")
    add_matrix_sheet("Von hoa", "MARKET_CAP_BN")
    add_matrix_sheet("GTGD_Von hoa", "GTGD_CAP_RATIO")

    sector_data = build_sector_flow_snapshot(years=years)
    ws_sector = wb.create_sheet(title="Nganh dashboard")
    ws_sector.append([
        "Rank", "Industry", "GTGD_BN", "Cap_BN", "GTGD_Cap_%", "Ty_trong_%",
        "GTGD_percentile", "GTGD_Cap_percentile", "Ty_trong_percentile",
        "GTGD_MA20", "GTGD_MA60", "Ty_trong_MA20", "Ty_trong_MA60",
        "Ty_trong_vs_MA60_pp"
    ])
    for item in sector_data.get("rows", []):
        ws_sector.append([
            item["Rank"], item["Industry"], item["GTGDBillion"], item["CapBillion"],
            item["GTGDCapPct"], item["MarketSharePct"], item["GTGDPercentile"],
            item["GTGDCapPercentile"], item["MarketSharePercentile"], item["GTGDMA20"],
            item["GTGDMA60"], item["ShareMA20"], item["ShareMA60"], item["ShareVsMA60Pct"]
        ])

    for cell in ws_sector[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=market_all_data_{max_date_str}_{years}y.xlsx"}
    )

# --- 4. MARKET ANALYSIS API ---
@app.get("/api/market-analysis")
def get_market_analysis():
    results = state.scan_state["results"]
    if not results:
        return {"breadth": {"ma20": 0.0, "ma50": 0.0, "ma100": 0.0, "ma200": 0.0}, "industries": [], "details": {}}
        
    df = pd.DataFrame(results)
    
    # Calculate global market breadth (exclude VNINDEX)
    df_stocks = df[df['Ticker'] != 'VNINDEX']
    total_stocks = len(df_stocks)
    
    breadth = {
        "ma20": float((df_stocks['MA20'] == 1).sum() / total_stocks if total_stocks > 0 else 0),
        "ma50": float((df_stocks['MA50'] == 1).sum() / total_stocks if total_stocks > 0 else 0),
        "ma100": float((df_stocks['MA100'] == 1).sum() / total_stocks if total_stocks > 0 else 0),
        "ma200": float((df_stocks['MA200'] == 1).sum() / total_stocks if total_stocks > 0 else 0)
    }
    
    # Retrieve VNINDEX return
    vnindex_row = df[df['Ticker'] == 'VNINDEX']
    vnindex_ret = float(vnindex_row['Return_2026_03_23'].iloc[0]) if not vnindex_row.empty else 0.0
    
    # Calculate industry metrics
    industries_list = []
    details_dict = {}
    
    for ind, group in df_stocks.groupby('Industry'):
        avg_ret = float(group['Return_2026_03_23'].mean())
        relative_ret = avg_ret - vnindex_ret
        
        industries_list.append({
            "industry": ind,
            "tickers_count": int(len(group)),
            "avg_return": avg_ret,
            "relative_to_vnindex": relative_ret
        })
        
        # Details of each ticker
        details_list = []
        for _, row in group.iterrows():
            details_list.append({
                "symbol": row['Ticker'],
                "return_pct": float(row['Return_2026_03_23'])
            })
        details_dict[ind] = details_list
        
    return {
        "breadth": breadth,
        "vnindex_return": vnindex_ret,
        "industries": industries_list,
        "details": details_dict
    }

# --- 5. ANTIGRAVITY SCANNER BACKGROUND TASK ---
def run_anti_worker():
    with state.anti_lock:
        state.anti_state["running"] = True
        state.anti_state["progress"] = 0.0
        state.anti_state["current"] = 0
        state.anti_state["total"] = len(SCAN_LIST)
        state.anti_state["signals"] = []
        state.anti_state["stats"] = {}
        state.anti_state["error"] = None
        
    total = len(SCAN_LIST)
    completed = 0
    signals_list = []
    signals_lock = threading.Lock()
    progress_lock = threading.Lock()
    
    q = queue.Queue()
    for symbol in SCAN_LIST:
        q.put(symbol)
        
    def worker():
        nonlocal completed
        while state.anti_state["running"]:
            try:
                symbol = q.get_nowait()
            except queue.Empty:
                break
                
            try:
                df = state.engine.get_history(symbol, start='2024-01-01')
                if df.empty or len(df) < 50:
                    q.task_done()
                    with progress_lock:
                        completed += 1
                        state.anti_state["current"] = completed
                        state.anti_state["progress"] = (completed / total) * 100
                    continue

                df = df.sort_values('time').reset_index(drop=True)
                df['vol_ma20'] = df['volume'].rolling(window=20).mean()
                df['vol_below_ma'] = df['volume'] < df['vol_ma20']
                df['consecutive_vol_low'] = df['vol_below_ma'].rolling(window=5).apply(lambda x: x.all(), raw=True)
                
                def is_sideways(window_closes):
                    if len(window_closes) < 5: return 0
                    p_min, p_max = window_closes.min(), window_closes.max()
                    return 1 if p_min > 0 and (p_max - p_min) / p_min <= 0.04 else 0

                df['sideways'] = df['close'].rolling(window=5).apply(is_sideways, raw=True)
                df['setup_met'] = (df['consecutive_vol_low'].shift(1) == 1) & (df['sideways'].shift(1) == 1)
                df['vol_spike'] = df['volume'] >= (1.5 * df['vol_ma20'])
                df['price_spike'] = (df['close'] / df['close'].shift(1) - 1) >= 0.03
                df['trigger'] = df['setup_met'] & df['vol_spike'] & df['price_spike']
                
                signals_idx = df[df['trigger'] == True].index.tolist()
                
                for idx in signals_idx:
                    row = df.iloc[idx]
                    s_idx = idx - 1
                    while s_idx > 0 and df.iloc[s_idx]['vol_below_ma']: 
                        s_idx -= 1
                    days_in_setup = idx - (s_idx + 1)
                    
                    def get_ret(n):
                        if idx + n < len(df): 
                            return (df.iloc[idx + n]['close'] / row['close'] - 1) * 100
                        return None
                    
                    ret5, ret10, ret20 = get_ret(5), get_ret(10), get_ret(20)
                    risk = row['close'] - df.iloc[idx-5:idx]['low'].min()
                    reward = df.iloc[idx:min(idx+21, len(df))]['high'].max() - row['close']
                    rr = reward / risk if risk > 0 else 0
                    
                    sig_data = {
                        "Ticker": symbol,
                        "Date": row['time'].strftime('%Y-%m-%d'),
                        "Days_Setup": int(days_in_setup),
                        "Return_5D": float(ret5) if ret5 is not None else 0.0,
                        "Return_10D": float(ret10) if ret10 is not None else 0.0,
                        "Return_20D": float(ret20) if ret20 is not None else 0.0,
                        "RR": float(rr)
                    }
                    
                    with signals_lock:
                        signals_list.append(sig_data)
                        
            except Exception as e:
                print(f"Error scanning Antigravity for {symbol}: {e}")
                
            q.task_done()
            with progress_lock:
                completed += 1
                state.anti_state["current"] = completed
                state.anti_state["progress"] = (completed / total) * 100

    num_workers = 8
    threads = []
    for _ in range(num_workers):
        t = threading.Thread(target=worker)
        t.start()
        threads.append(t)
        
    for t in threads:
        t.join()
        
    stats = {}
    if signals_list:
        rdf = pd.DataFrame(signals_list)
        stats = {
            "avg_days": float(rdf['Days_Setup'].mean()),
            "wr_5d": float((rdf['Return_5D'] > 0).mean()),
            "wr_10d": float((rdf['Return_10D'] > 0).mean()),
            "wr_20d": float((rdf['Return_20D'] > 0).mean()),
            "avg_rr": float(rdf['RR'].mean())
        }
    else:
        stats = {"avg_days": 0.0, "wr_5d": 0.0, "wr_10d": 0.0, "wr_20d": 0.0, "avg_rr": 0.0}

    with state.anti_lock:
        state.anti_state["signals"] = signals_list
        state.anti_state["stats"] = stats
        state.anti_state["progress"] = 100.0
        state.anti_state["running"] = False

@app.post("/api/antigravity/start")
def start_antigravity_scan(background_tasks: BackgroundTasks):
    if state.anti_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_anti_worker)
    return {"status": "started"}

@app.post("/api/antigravity/stop")
def stop_antigravity_scan():
    state.anti_state["running"] = False
    return {"status": "stopped"}

@app.get("/api/antigravity/status")
def get_antigravity_status():
    return {
        "running": state.anti_state["running"],
        "progress": state.anti_state["progress"],
        "current": state.anti_state["current"],
        "total": state.anti_state["total"]
    }

@app.get("/api/antigravity/results")
def get_antigravity_results():
    return {
        "signals": state.anti_state["signals"],
        "stats": state.anti_state["stats"]
    }

# --- 6. WATCHLIST SCANNER BACKGROUND TASK ---
def run_watch_worker():
    with state.watch_lock:
        state.watch_state["running"] = True
        state.watch_state["progress"] = 0.0
        state.watch_state["current"] = 0
        state.watch_state["total"] = len(SCAN_LIST)
        state.watch_state["watchlist"] = []
        state.watch_state["error"] = None
        
    total = len(SCAN_LIST)
    completed = 0
    watchlist_list = []
    watch_lock = threading.Lock()
    progress_lock = threading.Lock()
    
    q = queue.Queue()
    for symbol in SCAN_LIST:
        q.put(symbol)
        
    def worker():
        nonlocal completed
        while state.watch_state["running"]:
            try:
                symbol = q.get_nowait()
            except queue.Empty:
                break
                
            try:
                df = state.engine.get_history(symbol, length=50)
                if df.empty or len(df) < 25:
                    q.task_done()
                    with progress_lock:
                        completed += 1
                        state.watch_state["current"] = completed
                        state.watch_state["progress"] = (completed / total) * 100
                    continue

                df = df.sort_values('time').reset_index(drop=True)
                df['vol_ma20'] = df['volume'].rolling(window=20).mean()
                
                recent_5 = df.iloc[-5:]
                if recent_5['vol_ma20'].isnull().any():
                    q.task_done()
                    with progress_lock:
                        completed += 1
                        state.watch_state["current"] = completed
                        state.watch_state["progress"] = (completed / total) * 100
                    continue

                vol_below = (recent_5['volume'] < recent_5['vol_ma20']).all()
                if vol_below:
                    p_min = recent_5['low'].min()
                    p_max = recent_5['high'].max()
                    p_range_pct = (p_max - p_min) / p_min * 100 if p_min > 0 else 100
                    
                    if p_range_pct <= 6.0:
                        streak = 0
                        idx = len(df) - 1
                        while idx >= 0 and pd.notna(df.iloc[idx]['vol_ma20']) and \
                              df.iloc[idx]['volume'] < df.iloc[idx]['vol_ma20']:
                            streak += 1
                            idx -= 1
                        
                        current_row = df.iloc[-1]
                        
                        industry = "Index"
                        if symbol != 'VNINDEX':
                            for ind, symbols in VN302_INDUSTRIES.items():
                                if symbol in symbols:
                                    industry = ind
                                    break
                                    
                        watch_data = {
                            "Ticker": symbol,
                            "Price": float(current_row['close']),
                            "Consolidation_Days": int(streak),
                            "Vol_Ratio": float(current_row['volume'] / current_row['vol_ma20'] if current_row['vol_ma20'] > 0 else 0),
                            "Range_5D": float(p_range_pct),
                            "Industry": industry
                        }
                        
                        with watch_lock:
                            watchlist_list.append(watch_data)
                            
            except Exception as e:
                print(f"Error scanning Watchlist for {symbol}: {e}")
                
            q.task_done()
            with progress_lock:
                completed += 1
                state.watch_state["current"] = completed
                state.watch_state["progress"] = (completed / total) * 100

    num_workers = 8
    threads = []
    for _ in range(num_workers):
        t = threading.Thread(target=worker)
        t.start()
        threads.append(t)
        
    for t in threads:
        t.join()
        
    with state.watch_lock:
        state.watch_state["watchlist"] = watchlist_list
        state.watch_state["progress"] = 100.0
        state.watch_state["running"] = False

@app.post("/api/watchlist/start")
def start_watchlist_scan(background_tasks: BackgroundTasks):
    if state.watch_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_watch_worker)
    return {"status": "started"}

@app.post("/api/watchlist/stop")
def stop_watchlist_scan():
    state.watch_state["running"] = False
    return {"status": "stopped"}

@app.get("/api/watchlist/status")
def get_watchlist_status():
    return {
        "running": state.watch_state["running"],
        "progress": state.watch_state["progress"],
        "current": state.watch_state["current"],
        "total": state.watch_state["total"]
    }

@app.get("/api/watchlist/results")
def get_watchlist_results():
    return {"watchlist": state.watch_state["watchlist"]}

# --- 7. NEW: LIQUIDITY ("THANH KHOẢN") ENDPOINT ---
def init_db_liquidity():
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_liquidity (
                date TEXT,
                symbol TEXT,
                close REAL,
                volume INTEGER,
                liquidity_vnd INTEGER,
                industry TEXT,
                PRIMARY KEY (date, symbol)
            )
        """)
        conn.commit()
        conn.close()
        print("Database daily_liquidity cache table verified.")
    except Exception as e:
        print(f"Error initializing daily_liquidity cache table: {e}")

def normalize_price(price: float) -> float:
    if price is None:
        return 0.0
    return float(price) if float(price) > 1000 else float(price) * 1000


def get_symbol_industry(symbol: str) -> str:
    for ind, symbols in VN302_INDUSTRIES.items():
        if symbol in symbols:
            return ind
    return "Chưa phân loại"


def get_outstanding_shares(symbol: str, shares_map: Dict[str, int]) -> Optional[int]:
    shares = shares_map.get(symbol) or FALLBACK_SHARES.get(symbol)
    if shares and shares > 0:
        return int(shares)
    return None


def calculate_official_value(symbol: str, open_p: float, high_p: float, low_p: float, close_p: float, volume: int) -> int:
    """Calculates traded value as close price multiplied by matched volume."""
    close_vnd = normalize_price(close_p)
    return int(close_vnd * int(volume))


def run_liq_worker(date: str):
    from datetime import datetime, timedelta
    try:
        dt_obj = datetime.strptime(date, '%Y-%m-%d')
        start_dt_str = (dt_obj - timedelta(days=15)).strftime('%Y-%m-%d')
        df_vn = state.engine.get_history('VNINDEX', start=start_dt_str, end=date)
        if not df_vn.empty:
            date = df_vn['time'].dt.strftime('%Y-%m-%d').max()
    except Exception as e:
        print(f"Error finding nearest date: {e}")

    with state.liq_lock:
        state.liq_state["running"] = True
        state.liq_state["progress"] = 0.0
        state.liq_state["current"] = 0
        state.liq_state["total"] = len(VN302)
        state.liq_state["results"] = []
        state.liq_state["summary"] = {}
        state.liq_state["error"] = None
        state.liq_state["date"] = date
        
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        
        # Always aggregate from historical_prices so formula changes are reflected immediately.
        cursor.execute("SELECT symbol, open, high, low, close, volume FROM historical_prices WHERE time = ?", (date,))
        rows = cursor.fetchall()
        conn.close()
        
        results_list = []
        for r in rows:
            sym, o, h, l, c, v = r
            if sym in VN302 and v > 0:
                val_vnd = calculate_official_value(
                    symbol=sym, open_p=float(o), high_p=float(h),
                    low_p=float(l), close_p=float(c), volume=int(v)
                )
                industry = "Chưa phân loại"
                for ind, symbols in VN302_INDUSTRIES.items():
                    if sym in symbols:
                        industry = ind
                        break
                results_list.append({
                    "Ticker": sym,
                    "Industry": industry,
                    "Close": float(c) if c > 1000 else float(c * 1000),
                    "Volume": int(v),
                    "Liquidity_VND": val_vnd
                })
                
        if results_list:
            results_list = sorted(results_list, key=lambda x: x["Liquidity_VND"], reverse=True)
            total_value_vnd = sum(item["Liquidity_VND"] for item in results_list)
            total_volume = sum(item["Volume"] for item in results_list)
            leader_ticker = results_list[0]["Ticker"] if results_list else "N/A"
            avg_value_vnd = total_value_vnd / len(results_list) if results_list else 0
            
            summary = {
                "total_value_vnd": total_value_vnd,
                "total_volume": total_volume,
                "leader_ticker": leader_ticker,
                "avg_value_vnd": int(avg_value_vnd)
            }
            
            # Cache it to daily_liquidity for next time
            try:
                conn = get_db_conn()
                cursor = conn.cursor()
                for r in results_list:
                    cursor.execute("""
                        INSERT OR REPLACE INTO daily_liquidity (date, symbol, close, volume, liquidity_vnd, industry)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (date, r["Ticker"], r["Close"], r["Volume"], r["Liquidity_VND"], r["Industry"]))
                conn.commit()
                conn.close()
            except Exception as ce:
                print(f"Error caching compiled liquidity: {ce}")
                
            with state.liq_lock:
                state.liq_state["results"] = results_list
                state.liq_state["summary"] = summary
        else:
            with state.liq_lock:
                state.liq_state["results"] = []
                state.liq_state["summary"] = {"total_value_vnd": 0, "total_volume": 0, "leader_ticker": "N/A", "avg_value_vnd": 0}
                state.liq_state["error"] = "Dữ liệu ngày này chưa được đồng bộ. Vui lòng bấm nút 'Đồng bộ Data 3 năm' trước."
                
    except Exception as e:
        print(f"Error running liquidity worker: {e}")
        with state.liq_lock:
            state.liq_state["error"] = str(e)
    finally:
        with state.liq_lock:
            state.liq_state["progress"] = 100.0
            state.liq_state["running"] = False


# Background historical daemon crawler
def background_liquidity_crawler():
    print("Background Liquidity crawler daemon active.")
    # Allow 10 seconds for standard boot up
    time.sleep(10)
    
    try:
        # 1. Fetch valid trading sessions from benchmark index
        df_days = state.engine.get_history('VNINDEX', start='2026-01-16')
        if df_days.empty:
            print("Crawler warning: VNINDEX history is empty. Delaying crawl.")
            return
            
        trading_dates = sorted(df_days['time'].dt.strftime('%Y-%m-%d').unique().tolist())
        print(f"Crawler identified {len(trading_dates)} trading sessions since 2026-01-16.")
        
        # Check already cached dates
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT date FROM daily_liquidity")
        cached_dates = {row[0] for row in cursor.fetchall()}
        conn.close()
        
        missing_dates = [d for d in trading_dates if d not in cached_dates]
        print(f"Crawler will crawl and cache {len(missing_dates)} missing dates.")
        
        for date in missing_dates:
            print(f"Crawler pre-fetching and caching date: {date}...")
            # Run quick crawler scan for this date
            results_list = []
            for symbol in VN302:
                df = state.engine.get_history(symbol, start=date, end=date)
                if not df.empty:
                    row = df.iloc[-1]
                    vol = int(row['volume'])
                    if vol > 0:
                        val_vnd = calculate_official_value(
                            symbol=symbol,
                            open_p=float(row['open']),
                            high_p=float(row['high']),
                            low_p=float(row['low']),
                            close_p=float(row['close']),
                            volume=vol
                        )
                        industry = "Chưa phân loại"
                        for ind, symbols in VN302_INDUSTRIES.items():
                            if symbol in symbols:
                                industry = ind
                                break
                        results_list.append({
                            "Ticker": symbol,
                            "Industry": industry,
                            "Close": float(row['close']) if row['close'] > 1000 else float(row['close'] * 1000),
                            "Volume": vol,
                            "Liquidity_VND": val_vnd
                        })
            
            if results_list:
                try:
                    conn = get_db_conn()
                    cursor = conn.cursor()
                    for r in results_list:
                        cursor.execute("""
                            INSERT OR REPLACE INTO daily_liquidity (date, symbol, close, volume, liquidity_vnd, industry)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (date, r["Ticker"], r["Close"], r["Volume"], r["Liquidity_VND"], r["Industry"]))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    print(f"Crawler error saving date {date}: {e}")
                    
            # Be gentle with a 1.5 second wait between sessions
            time.sleep(1.5)
            
        print("Crawler finished caching all missing historical liquidity data.")
        
        # Periodic update loop while the web server is open (once per day).
        while True:
            time.sleep(24 * 3600)
            print("Crawler periodic daily check started...")
            df_days = state.engine.get_history('VNINDEX', start='2026-01-16')
            if not df_days.empty:
                trading_dates = sorted(df_days['time'].dt.strftime('%Y-%m-%d').unique().tolist())
                conn = get_db_conn()
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT date FROM daily_liquidity")
                cached_dates = {row[0] for row in cursor.fetchall()}
                conn.close()
                
                missing_dates = [d for d in trading_dates if d not in cached_dates]
                for date in missing_dates:
                    print(f"Crawler found new session date: {date}. Caching...")
                    results_list = []
                    for symbol in VN302:
                        df = state.engine.get_history(symbol, start=date, end=date)
                        if not df.empty:
                            row = df.iloc[-1]
                            vol = int(row['volume'])
                            if vol > 0:
                                val_vnd = calculate_official_value(
                                    symbol=symbol,
                                    open_p=float(row['open']),
                                    high_p=float(row['high']),
                                    low_p=float(row['low']),
                                    close_p=float(row['close']),
                                    volume=vol
                                )
                                industry = "Chưa phân loại"
                                for ind, symbols in VN302_INDUSTRIES.items():
                                    if symbol in symbols:
                                        industry = ind
                                        break
                                results_list.append({
                                    "Ticker": symbol,
                                    "Industry": industry,
                                    "Close": float(row['close']) if row['close'] > 1000 else float(row['close'] * 1000),
                                    "Volume": vol,
                                    "Liquidity_VND": val_vnd
                                })
                    if results_list:
                        try:
                            conn = get_db_conn()
                            cursor = conn.cursor()
                            for r in results_list:
                                cursor.execute("""
                                    INSERT OR REPLACE INTO daily_liquidity (date, symbol, close, volume, liquidity_vnd, industry)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (date, r["Ticker"], r["Close"], r["Volume"], r["Liquidity_VND"], r["Industry"]))
                            conn.commit()
                            conn.close()
                        except Exception as e:
                            print(f"Crawler periodic save error for date {date}: {e}")
                            
    except Exception as e:
        print(f"Crawler crashed: {e}")

def start_background_crawler():
    t = threading.Thread(target=background_liquidity_crawler, daemon=True)
    t.start()

@app.post("/api/liquidity/start")
def start_liquidity_scan(background_tasks: BackgroundTasks, date: str = Query('2026-05-18')):
    if state.liq_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_liq_worker, date)
    return {"status": "started"}

@app.post("/api/liquidity/stop")
def stop_liquidity_scan():
    state.liq_state["running"] = False
    return {"status": "stopped"}

@app.get("/api/liquidity/status")
def get_liquidity_status():
    return {
        "running": state.liq_state["running"],
        "progress": state.liq_state["progress"],
        "current": state.liq_state["current"],
        "total": state.liq_state["total"],
        "date": state.liq_state["date"]
    }

@app.get("/api/liquidity/results")
def get_liquidity_results():
    return {
        "summary": state.liq_state["summary"],
        "liquidity": state.liq_state["results"]
    }

# --- 7.5 NEW: MARKET CAPITALIZATION ("VỐN HÓA") SYSTEM ---



FALLBACK_SHARES = {
    'VIC': 7706031024,
    'HPG': 7675465855,
    'VCB': 5589062006,
    'VHM': 4354367468,
    'VPB': 7933923485,
    'TCB': 7044810714,
    'BID': 5700435900,
    'MBB': 5287093244,
    'CTG': 5370072931,
    'ACB': 4466276856,
    'VRE': 2272318410,
    'FPT': 1460492823,
    'GAS': 2296720000,
    'MSN': 1430843406,
    'VNM': 2089955445,
    'SSI': 1511130182,
    'VND': 1222044101,
    'DGC': 379769363,
    'HSG': 615931652,
    'NKG': 263353815,
    'PVD': 555831511,
    'PVS': 477966290,
}

def init_db_shares():
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ticker_shares (
                symbol TEXT PRIMARY KEY,
                outstanding_shares INTEGER
            )
        """)
        conn.commit()
        # Seed the up-to-date fallback ones
        for sym, sh in FALLBACK_SHARES.items():
            cursor.execute("INSERT OR IGNORE INTO ticker_shares (symbol, outstanding_shares) VALUES (?, ?)", (sym, sh))
        conn.commit()
        conn.close()
        print("Database ticker_shares cache table verified and seeded with up-to-date data.")
    except Exception as e:
        print(f"Error initializing ticker_shares cache table: {e}")


@app.post("/api/shares-crawler/restart")
def restart_shares_crawler(background_tasks: BackgroundTasks):
    if state.shares_crawler_state["running"]:
        return {"status": "already_running"}
        
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM ticker_shares")
        conn.commit()
        # Seed the up-to-date fallback ones
        for sym, sh in FALLBACK_SHARES.items():
            cursor.execute("INSERT OR IGNORE INTO ticker_shares (symbol, outstanding_shares) VALUES (?, ?)", (sym, sh))
        conn.commit()
        conn.close()
        print("Shares database cache wiped for fresh re-crawl.")
    except Exception as e:
        print(f"Error wiping shares cache: {e}")
        
    background_tasks.add_task(background_shares_crawler)
    return {"status": "started"}


def background_shares_crawler():
    print("Background outstanding shares crawler active.")
    with state.shares_crawler_lock:
        state.shares_crawler_state["running"] = True
        state.shares_crawler_state["completed"] = False
        state.shares_crawler_state["total"] = len(VN302)
        
    time.sleep(10) # Wait for web server boot to settle
    
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT symbol FROM ticker_shares")
        cached_syms = {row[0] for row in cursor.fetchall()}
        conn.close()
        
        total = len(VN302)
        missing_syms = [s for s in VN302 if s not in cached_syms]
        print(f"Shares crawler: {len(missing_syms)} missing symbols to fetch.")
        
        current = len(cached_syms)
        with state.shares_crawler_lock:
            state.shares_crawler_state["current"] = current
            state.shares_crawler_state["progress"] = round((current / total) * 100.0, 1)
            if current >= total:
                state.shares_crawler_state["completed"] = True
                state.shares_crawler_state["running"] = False
                
        for symbol in missing_syms:
            try:
                c = Company(symbol=symbol, source="KBS")
                df = c.overview()
                if not df.empty and 'outstanding_shares' in df.columns:
                    shares = int(df['outstanding_shares'].iloc[0])
                    if shares > 0:
                        conn = get_db_conn()
                        cursor = conn.cursor()
                        cursor.execute("INSERT OR REPLACE INTO ticker_shares (symbol, outstanding_shares) VALUES (?, ?)", (symbol, shares))
                        conn.commit()
                        conn.close()
                        print(f"Fetched & cached shares for {symbol}: {shares:,}")
            except Exception as e:
                pass
            
            current += 1
            with state.shares_crawler_lock:
                state.shares_crawler_state["current"] = current
                state.shares_crawler_state["progress"] = round((current / total) * 100.0, 1)
                
            time.sleep(2.5) # Gentle throttling to stay under 60 req/min
            
        with state.shares_crawler_lock:
            state.shares_crawler_state["completed"] = True
            state.shares_crawler_state["progress"] = 100.0
            print("Background outstanding shares crawler completed successfully!")
            
    except Exception as e:
        print(f"Error in shares crawler: {e}")
    finally:
        with state.shares_crawler_lock:
            state.shares_crawler_state["running"] = False

def start_shares_crawler():
    t = threading.Thread(target=background_shares_crawler, daemon=True)
    t.start()


def run_market_cap_worker(date: str):
    from datetime import datetime, timedelta
    try:
        dt_obj = datetime.strptime(date, '%Y-%m-%d')
        start_dt_str = (dt_obj - timedelta(days=15)).strftime('%Y-%m-%d')
        df_vn = state.engine.get_history('VNINDEX', start=start_dt_str, end=date)
        if not df_vn.empty:
            date = df_vn['time'].dt.strftime('%Y-%m-%d').max()
    except Exception as e:
        print(f"Error finding nearest date: {e}")

    with state.cap_lock:
        state.cap_state["running"] = True
        state.cap_state["progress"] = 0.0
        state.cap_state["current"] = 0
        state.cap_state["total"] = len(VN302)
        state.cap_state["results"] = []
        state.cap_state["industries_summary"] = []
        state.cap_state["top_10"] = []
        state.cap_state["error"] = None
        state.cap_state["date"] = date

    try:
        shares_map = {}
        try:
            conn = get_db_conn()
            cursor = conn.cursor()
            cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
            for row in cursor.fetchall():
                shares_map[row[0]] = row[1]
            conn.close()
        except Exception as e:
            print(f"Error loading ticker_shares: {e}")



        # Fetch from historical_prices instantly (NO API calls!)
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT symbol, close, volume FROM historical_prices WHERE time = ?", (date,))
        rows = cursor.fetchall()
        conn.close()
        
        price_map = {r[0]: r[1] for r in rows}
        volume_map = {r[0]: r[2] for r in rows}
        results_list = []
        
        for symbol in VN302:
            close_val = price_map.get(symbol, 0.0)
            volume_val = volume_map.get(symbol)
            shares = get_outstanding_shares(symbol, shares_map)
                
            if close_val > 0 and shares:
                actual_price = normalize_price(close_val)
                market_cap_billion = round((actual_price * shares) / 1_000_000_000, 2)
                liquidity_billion = None
                if volume_val is not None:
                    liquidity_billion = round((actual_price * int(volume_val)) / 1_000_000_000, 2)
                
                industry = "Chưa phân loại"
                for ind, symbols in VN302_INDUSTRIES.items():
                    if symbol in symbols:
                        industry = ind
                        break
                        
                results_list.append({
                    "Ticker": symbol,
                    "Industry": industry,
                    "Close": actual_price,
                    "Volume": int(volume_val) if volume_val is not None else None,
                    "OutstandingShares": shares,
                    "MarketCapBillion": market_cap_billion,
                    "LiquidityBillion": liquidity_billion
                })
                
        if results_list:
            results_list = sorted(results_list, key=lambda x: x["MarketCapBillion"], reverse=True)
            for rank, r in enumerate(results_list, 1):
                r["Rank"] = rank
                
            top_10 = results_list[:10]
            
            industry_data = {}
            for r in results_list:
                ind = r["Industry"]
                if ind not in industry_data:
                    industry_data[ind] = {"total_cap": 0.0, "count": 0}
                industry_data[ind]["total_cap"] += r["MarketCapBillion"]
                industry_data[ind]["count"] += 1
                
            industry_summary = []
            for ind, data in industry_data.items():
                industry_summary.append({
                    "Industry": ind,
                    "TotalCapBillion": round(data["total_cap"], 2),
                    "Count": data["count"]
                })
            industry_summary = sorted(industry_summary, key=lambda x: x["TotalCapBillion"], reverse=True)
            
            with state.cap_lock:
                state.cap_state["results"] = results_list
                state.cap_state["top_10"] = top_10
                state.cap_state["industries_summary"] = industry_summary
        else:
            with state.cap_lock:
                state.cap_state["error"] = "Dữ liệu ngày này chưa được đồng bộ. Vui lòng bấm nút 'Đồng bộ Data 3 năm' trước."
                
    except Exception as e:
        print(f"Error running market cap worker: {e}")
        with state.cap_lock:
            state.cap_state["error"] = str(e)
    finally:
        with state.cap_lock:
            state.cap_state["progress"] = 100.0
            state.cap_state["running"] = False


@app.post("/api/market-cap/start")
def start_market_cap_scan(background_tasks: BackgroundTasks, date: str = Query('2026-05-18')):
    if state.cap_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_market_cap_worker, date)
    return {"status": "started"}

@app.post("/api/market-cap/stop")
def stop_market_cap_scan():
    state.cap_state["running"] = False
    return {"status": "stopped"}

@app.get("/api/market-cap/status")
def get_market_cap_status():
    return {
        "running": state.cap_state["running"],
        "progress": state.cap_state["progress"],
        "current": state.cap_state["current"],
        "total": state.cap_state["total"],
        "date": state.cap_state["date"]
    }

@app.get("/api/market-cap/results")
def get_market_cap_results():
    return {
        "results": state.cap_state["results"],
        "top_10": state.cap_state["top_10"],
        "industries_summary": state.cap_state["industries_summary"]
    }

@app.get("/api/export/market-cap")
def export_market_cap_excel(date: str = Query('2026-05-18')):
    results = state.cap_state["results"]
    if not results:
        raise HTTPException(status_code=400, detail="No market cap data available. Please run scan first.")
        
    df_results = pd.DataFrame(results)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name="Chi Tiết Vốn Hóa", index=False)
        if state.cap_state["industries_summary"]:
            df_ind = pd.DataFrame(state.cap_state["industries_summary"])
            df_ind.to_excel(writer, sheet_name="Tổng Hợp Ngành", index=False)
            
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=market_cap_{date}_report.xlsx"}
    )

# --- 7.7 NEW: MARKET CAP TIME RANGE DAEMON & ROUTES ---


def run_sync_historical_worker():
    with state.sync_lock:
        state.sync_state["running"] = True
        state.sync_state["progress"] = 0.0
        state.sync_state["current"] = 0
        state.sync_state["total"] = len(VN302)
        state.sync_state["error"] = None

    try:
        # Step 1: Download/update price history from 2023-01-01 to today for all 302 symbols
        total = len(VN302)
        completed = 0
        progress_lock = threading.Lock()
        q = queue.Queue()
        for symbol in VN302:
            q.put(symbol)

        def worker():
            nonlocal completed
            while state.sync_state["running"]:
                try:
                    symbol = q.get_nowait()
                except queue.Empty:
                    break
                try:
                    # Sync 3 years of data (saves into historical_prices)
                    state.engine.get_history(symbol, start='2023-01-01')
                except Exception as e:
                    print(f"Error syncing {symbol}: {e}")
                finally:
                    q.task_done()
                    with progress_lock:
                        completed += 1
                        state.sync_state["current"] = completed
                        state.sync_state["progress"] = (completed / total) * 80

        threads = []
        for _ in range(8):
            t = threading.Thread(target=worker)
            t.start()
            threads.append(t)
        
        for t in threads:
            t.join()

        # Step 2: Post-process daily_liquidity cache for ALL dates from 2023-01-01 to today
        if state.sync_state["running"]:
            print("Sync complete. Aggregating historical liquidity records...")
            conn = get_db_conn()
            cursor = conn.cursor()
            
            cursor.execute("SELECT symbol, time, open, high, low, close, volume FROM historical_prices WHERE time >= '2023-01-01'")
            rows = cursor.fetchall()
            
            prices_by_date = {}
            for r in rows:
                sym, t, o, h, l, c, v = r
                if t not in prices_by_date:
                    prices_by_date[t] = []
                prices_by_date[t].append((sym, o, h, l, c, v))
                
            total_dates = len(prices_by_date)
            date_idx = 0
            
            for date_str, symbol_rows in prices_by_date.items():
                if not state.sync_state["running"]:
                    break
                
                insert_batch = []
                for sym, o, h, l, c, v in symbol_rows:
                    if sym in VN302 and v > 0:
                        val_vnd = calculate_official_value(
                            symbol=sym, open_p=float(o), high_p=float(h),
                            low_p=float(l), close_p=float(c), volume=int(v)
                        )
                        industry = "Chưa phân loại"
                        for ind, symbols in VN302_INDUSTRIES.items():
                            if sym in symbols:
                                industry = ind
                                break
                        insert_batch.append((date_str, sym, float(c) if c > 1000 else float(c * 1000), int(v), val_vnd, industry))
                
                if insert_batch:
                    cursor.executemany("""
                        INSERT OR REPLACE INTO daily_liquidity (date, symbol, close, volume, liquidity_vnd, industry)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, insert_batch)
                    conn.commit()
                
                date_idx += 1
                state.sync_state["progress"] = 80 + (date_idx / total_dates) * 20
                
            conn.close()
            print("Finished historical liquidity aggregation!")
            
    except Exception as e:
        print(f"Error in sync worker: {e}")
        with state.sync_lock:
            state.sync_state["error"] = str(e)
    finally:
        with state.sync_lock:
            state.sync_state["progress"] = 100.0
            state.sync_state["running"] = False


@app.post("/api/sync/start")
def start_sync_scan(background_tasks: BackgroundTasks):
    if state.sync_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_sync_historical_worker)
    if not state.shares_crawler_state["running"]:
        background_tasks.add_task(background_shares_crawler)
    return {"status": "started"}

@app.post("/api/sync/stop")
def stop_sync_scan():
    state.sync_state["running"] = False
    return {"status": "stopped"}

@app.get("/api/sync/status")
def get_sync_status():
    return {
        "running": state.sync_state["running"],
        "progress": state.sync_state["progress"],
        "current": state.sync_state["current"],
        "total": state.sync_state["total"],
        "error": state.sync_state["error"]
    }

# --- 8. HISTORICAL PRICE NATIVE MODAL CHART DATA ---
@app.get("/api/history/{symbol}")
def get_ticker_history(symbol: str, source: str = 'KBS'):
    symbol = symbol.upper()
    try:
        df = state.engine.get_history(symbol, start='2025-06-01')
        if df.empty:
            return {"error": "Empty history dataframe."}
            
        df = df.sort_values('time').reset_index(drop=True)
        df['ma50'] = df['close'].rolling(50).mean()
        
        history_list = []
        for _, row in df.iterrows():
            close_p = float(row['close'])
            close_vnd = close_p if close_p > 1000 else close_p * 1000
            
            ma50_val = row['ma50']
            ma50_vnd = None
            if pd.notna(ma50_val):
                ma50_vnd = float(ma50_val) if ma50_val > 1000 else float(ma50_val * 1000)
                
            history_list.append({
                "date": row['time'].strftime('%Y-%m-%d'),
                "close": close_vnd,
                "ma50": ma50_vnd
            })
            
        return {
            "symbol": symbol,
            "history": history_list
        }
    except Exception as e:
        return {"error": str(e)}

# --- 9. EXCEL AND CSV DYNAMIC DOWNLOAD GENERATORS ---
@app.get("/api/export/scan")
def export_scan_csv():
    results = state.scan_state["results"]
    if not results:
        raise HTTPException(status_code=400, detail="No scan data available. Run scan first.")
        
    df = pd.DataFrame(results)
    stream = io.StringIO()
    df.to_csv(stream, index=False, encoding='utf-8-sig')
    
    response = StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=vn302_scanner_report.csv"}
    )
    return response

@app.get("/api/export/market")
def export_market_excel(source: str = 'KBS'):
    # Simulates the PyQt export market method generating multi-sheet excels
    results = state.scan_state["results"]
    if not results:
        raise HTTPException(status_code=400, detail="No data available. Run scanner first.")
        
    df_scan = pd.DataFrame(results)
    df_stocks = df_scan[df_scan['Ticker'] != 'VNINDEX']
    
    vnindex_row = df_scan[df_scan['Ticker'] == 'VNINDEX']
    vnindex_ret = float(vnindex_row['Return_2026_03_23'].iloc[0]) if not vnindex_row.empty else 0.0
    
    # 1. Industry Summary Sheet
    industry_summary = []
    for ind, group in df_stocks.groupby('Industry'):
        avg_ret = float(group['Return_2026_03_23'].mean())
        industry_summary.append({
            "Industry": ind,
            "Tickers Count": len(group),
            "Avg Return (%)": round(avg_ret, 2),
            "VNINDEX Return (%)": round(vnindex_ret, 2),
            "Relative Return (%)": round(avg_ret - vnindex_ret, 2),
            "Status": "Outperform" if (avg_ret - vnindex_ret) >= 0 else "Underperform"
        })
    df_ind = pd.DataFrame(industry_summary)
    
    # 2. Detailed Sheet
    df_det = df_stocks[['Ticker', 'Industry', 'Price', 'Return_2026_03_23', 'RSI', 'MACD', 'Breakout']].copy()
    df_det.columns = ['Ticker', 'Industry', 'Close Price', 'Return Since 2026-03-23 (%)', 'RSI (14)', 'MACD Status', '52w Breakout']
    
    # Write to Excel in memory using BytesIO
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_ind.to_excel(writer, sheet_name="Industry Summary", index=False)
        df_det.to_excel(writer, sheet_name="Ticker Details", index=False)
        
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=market_analysis_report_{source}.xlsx"}
    )

@app.get("/api/export/antigravity")
def export_antigravity_excel():
    signals = state.anti_state["signals"]
    stats = state.anti_state["stats"]
    
    if not signals:
        raise HTTPException(status_code=400, detail="No antigravity data to export.")
        
    df_sig = pd.DataFrame(signals)
    df_stats = pd.DataFrame([stats])
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_sig.to_excel(writer, sheet_name="Antigravity Signals", index=False)
        df_stats.to_excel(writer, sheet_name="Win-Rates Metrics", index=False)
        
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=antigravity_signals_report.xlsx"}
    )

@app.get("/api/export/watchlist")
def export_watchlist_excel():
    watchlist = state.watch_state["watchlist"]
    if not watchlist:
        raise HTTPException(status_code=400, detail="No watchlist data to export.")
        
    df_w = pd.DataFrame(watchlist)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_w.to_excel(writer, sheet_name="Watchlist Setup", index=False)
        
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=watchlist_report.xlsx"}
    )

@app.get("/api/export/backtest")
def export_backtest_excel(
    symbol: str, start_date: str, end_date: str, initial_capital: float,
    position_size: float, rsi_period: int, buy_threshold: float, sell_threshold: float,
    stop_loss: str, take_profit: str
):
    try:
        sl_val = stop_loss.strip().lower()
        sl = float(sl_val) if sl_val != 'none' and sl_val != '' else None
        
        tp_val = take_profit.strip().lower()
        tp = float(tp_val) if tp_val != 'none' and tp_val != '' else None
        
        bt = Backtester(
            symbol=symbol.upper(),
            start_date=start_date,
            end_date=end_date,
            initial_capital=initial_capital,
            rsi_period=rsi_period,
            buy_threshold=buy_threshold,
            sell_threshold=sell_threshold,
            stop_loss=sl,
            take_profit=tp,
            position_size=position_size,
            engine=state.engine
        )
        
        trades, metrics, df_curves = bt.run_backtest()
        
        # Write backtest sheets
        df_trades = pd.DataFrame(trades)
        if not df_trades.empty:
            df_trades['entry_date'] = df_trades['entry_date'].dt.strftime('%Y-%m-%d')
            df_trades['exit_date'] = df_trades['exit_date'].dt.strftime('%Y-%m-%d')
            
        df_metrics = pd.DataFrame([metrics])
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_trades.to_excel(writer, sheet_name="Trade Logs", index=False)
            df_metrics.to_excel(writer, sheet_name="Performance Metrics", index=False)
            if df_curves is not None and not df_curves.empty:
                df_c_write = df_curves.copy()
                df_c_write['time'] = df_c_write['time'].dt.strftime('%Y-%m-%d')
                df_c_write.to_excel(writer, sheet_name="Equity Curve", index=False)
                
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=backtest_{symbol.upper()}_report.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- 7.8 NEW: MARKET CAP RANGE CALCULATOR & EXCEL GROUPED EXPORTER ---

def run_market_cap_range_worker(start_date: str, end_date: str):
    with state.cap_range_lock:
        state.cap_range_state["running"] = True
        state.cap_range_state["progress"] = 0.0
        state.cap_range_state["current"] = 0
        state.cap_range_state["total"] = 0
        state.cap_range_state["results"] = []
        state.cap_range_state["details_by_date"] = {}
        state.cap_range_state["error"] = None

    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        
        # Get unique trading dates in range
        cursor.execute("""
            SELECT DISTINCT time FROM historical_prices 
            WHERE time >= ? AND time <= ? 
            ORDER BY time DESC
        """, (start_date, end_date))
        dates = [row[0] for row in cursor.fetchall()]
        
        if not dates:
            raise Exception("Không tìm thấy dữ liệu giao dịch trong khoảng thời gian này. Vui lòng bấm 'Đồng bộ Data 3 năm' trước.")
            
        total_days = len(dates)
        with state.cap_range_lock:
            state.cap_range_state["total"] = total_days
            
        # Get outstanding shares
        cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
        shares_map = {row[0]: row[1] for row in cursor.fetchall()}
        
        # Pull historical prices
        cursor.execute("""
            SELECT symbol, time, close, volume FROM historical_prices 
            WHERE time >= ? AND time <= ?
        """, (start_date, end_date))
        rows = cursor.fetchall()
        conn.close()
        
        price_matrix = {}
        volume_matrix = {}
        for sym, dt, cl, vol in rows:
            price_matrix.setdefault(dt, {})[sym] = cl
            volume_matrix.setdefault(dt, {})[sym] = vol
            
        results = []
        details_by_date = {}
        for i, dt in enumerate(dates):
            day_prices = price_matrix.get(dt, {})
            day_volumes = volume_matrix.get(dt, {})
            details = []
            
            for sym in VN302:
                cl = day_prices.get(sym)
                shares = get_outstanding_shares(sym, shares_map)
                if cl is None or cl <= 0 or not shares:
                    continue
                
                vol = day_volumes.get(sym)
                actual_price = normalize_price(cl)
                market_cap_billion = round((actual_price * shares) / 1_000_000_000, 2)
                liquidity_billion = None
                if vol is not None:
                    liquidity_billion = round((actual_price * int(vol)) / 1_000_000_000, 2)
                
                details.append({
                    "Ticker": sym,
                    "Industry": get_symbol_industry(sym),
                    "Close": actual_price,
                    "Volume": int(vol) if vol is not None else None,
                    "OutstandingShares": shares,
                    "MarketCapBillion": market_cap_billion,
                    "LiquidityBillion": liquidity_billion
                })

            details = sorted(details, key=lambda x: x["MarketCapBillion"], reverse=True)
            for rank, item in enumerate(details, 1):
                item["Rank"] = rank

            liq_details = [item for item in details if item["LiquidityBillion"] is not None]
            cap_leader = details[0] if details else None
            liq_leader = max(liq_details, key=lambda x: x["LiquidityBillion"], default=None)
            total_cap = sum(item["MarketCapBillion"] for item in details)
            total_liq = sum(item["LiquidityBillion"] for item in liq_details)
            details_by_date[dt] = details
                    
            results.append({
                "Date": dt,
                "date": dt,
                "TotalCapBillion": round(total_cap, 2),
                "TotalLiquidityBillion": round(total_liq, 2),
                "LeaderTicker": cap_leader["Ticker"] if cap_leader else "N/A",
                "LeaderCapBillion": cap_leader["MarketCapBillion"] if cap_leader else 0,
                "LiquidityLeaderTicker": liq_leader["Ticker"] if liq_leader else "N/A",
                "LiquidityLeaderBillion": liq_leader["LiquidityBillion"] if liq_leader else 0,
                "ValidCount": len(details)
            })
            
            with state.cap_range_lock:
                state.cap_range_state["current"] = i + 1
                state.cap_range_state["progress"] = round(((i + 1) / total_days) * 100.0, 1)
                
        with state.cap_range_lock:
            state.cap_range_state["results"] = results
            state.cap_range_state["details_by_date"] = details_by_date
            
    except Exception as e:
        print(f"Error in cap range worker: {e}")
        with state.cap_range_lock:
            state.cap_range_state["error"] = str(e)
    finally:
        with state.cap_range_lock:
            state.cap_range_state["running"] = False
            state.cap_range_state["progress"] = 100.0

@app.post("/api/market-cap-range/start")
def start_market_cap_range_scan(background_tasks: BackgroundTasks, start_date: str = Query(...), end_date: str = Query(...)):
    if state.cap_range_state["running"]:
        return {"status": "already_running"}
    background_tasks.add_task(run_market_cap_range_worker, start_date, end_date)
    return {"status": "started"}

@app.get("/api/market-cap-range/status")
def get_market_cap_range_status():
    return {
        "running": state.cap_range_state["running"],
        "progress": state.cap_range_state["progress"],
        "current": state.cap_range_state["current"],
        "total": state.cap_range_state["total"],
        "error": state.cap_range_state["error"]
    }


@app.get("/api/shares-crawler/status")
def get_shares_crawler_status():
    return {
        "running": state.shares_crawler_state["running"],
        "progress": state.shares_crawler_state["progress"],
        "current": state.shares_crawler_state["current"],
        "total": state.shares_crawler_state["total"],
        "completed": state.shares_crawler_state["completed"]
    }


@app.get("/api/market-cap-range/results")
def get_market_cap_range_results():
    results = state.cap_range_state["results"]
    error = state.cap_range_state["error"]
    details_by_date = state.cap_range_state.get("details_by_date", {})
    latest_results = details_by_date.get(results[0]["Date"], []) if results else []
    return {
        "daily_summaries": results,
        "latest_results": latest_results,
        "details_by_date": details_by_date,
        "error": error
    }
    
    latest_results = []
    if results:
        latest_date = results[0]["date"]
        try:
            conn = get_db_conn()
            cursor = conn.cursor()
            cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
            shares_map = {row[0]: row[1] for row in cursor.fetchall()}
            cursor.execute("SELECT symbol, close FROM historical_prices WHERE time = ?", (latest_date,))
            rows = cursor.fetchall()
            conn.close()
            
            price_map = {r[0]: r[1] for r in rows}
            temp_results = []
            for symbol in VN302:
                close_val = price_map.get(symbol, 0.0)
                shares = shares_map.get(symbol, 100_000_000)
                if close_val > 0:
                    actual_price = close_val if close_val > 1000 else close_val * 1000
                    market_cap_billion = round((actual_price * shares) / 1_000_000_000, 2)
                    
                    industry = "Chưa phân loại"
                    for ind, symbols in VN302_INDUSTRIES.items():
                        if symbol in symbols:
                            industry = ind
                            break
                            
                    temp_results.append({
                        "Ticker": symbol,
                        "Industry": industry,
                        "Close": actual_price,
                        "OutstandingShares": shares,
                        "MarketCapBillion": market_cap_billion
                    })
            if temp_results:
                temp_results = sorted(temp_results, key=lambda x: x["MarketCapBillion"], reverse=True)
                for rank, r in enumerate(temp_results, 1):
                    r["Rank"] = rank
                latest_results = temp_results
        except Exception as e:
            print(f"Error fetching latest day details: {e}")
            
    return {
        "daily_summaries": results,
        "latest_results": latest_results,
        "error": error
    }

@app.get("/api/export/market-cap-range")
def export_market_cap_range_excel(start_date: str = Query(...), end_date: str = Query(...)):
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        
        # 1. Fetch trade dates
        cursor.execute("""
            SELECT DISTINCT time FROM historical_prices 
            WHERE time >= ? AND time <= ? 
            ORDER BY time DESC
        """, (start_date, end_date))
        dates = [row[0] for row in cursor.fetchall()]
        
        if not dates:
            raise HTTPException(status_code=400, detail="Không có dữ liệu trong khoảng thời gian này. Vui lòng bấm 'Đồng bộ Data 3 năm' trước.")
            
        # 2. Fetch outstanding shares
        cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
        shares_map = {row[0]: row[1] for row in cursor.fetchall()}
        
        # 3. Fetch historical prices & volumes
        cursor.execute("""
            SELECT symbol, time, close, volume FROM historical_prices 
            WHERE time >= ? AND time <= ?
        """, (start_date, end_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Matrix mappings
        price_matrix = {}
        volume_matrix = {}
        for r in rows:
            sym, dt, cl, vol = r
            if dt not in price_matrix:
                price_matrix[dt] = {}
                volume_matrix[dt] = {}
            price_matrix[dt][sym] = cl
            volume_matrix[dt][sym] = vol
            
        # Group tickers by Industry horizontally
        ordered_tickers = []
        ticker_to_industry = {}
        for ind, syms in VN302_INDUSTRIES.items():
            for sym in syms:
                if sym in VN302:
                    ordered_tickers.append(sym)
                    ticker_to_industry[sym] = ind
        for sym in VN302:
            if sym not in ticker_to_industry:
                ordered_tickers.append(sym)
                ticker_to_industry[sym] = "Các ngành khác"
                
        # Generate Excel
        wb = Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Styles
        font_industry = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_ticker = Font(name="Segoe UI", size=10, bold=True)
        font_subheader = Font(name="Segoe UI", size=9, italic=True, color="555555")
        font_data = Font(name="Segoe UI", size=10)
        font_dates = Font(name="Segoe UI", size=10, bold=True)
        
        industry_colors = {
            "Ngân hàng": "366092",
            "Bất động sản": "E46C0A",
            "BĐS": "E46C0A",
            "Chứng khoán": "76933C",
            "Dầu khí": "953734",
            "Thép": "60497A",
            "Hàng không": "31859C",
            "Công nghệ": "7030A0",
            "Năng lượng": "595959",
            "Bán lẻ": "B9CDE5",
            "Xây dựng & Vật liệu": "D8E4BC",
            "Các ngành khác": "7F7F7F"
        }
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        def populate_worksheet(ws, metric_type):
            ws.cell(row=3, column=1, value="Dates").font = font_dates
            ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=1).border = thin_border
            
            # Write row 1, 2, 3
            for col_idx, sym in enumerate(ordered_tickers, 2):
                ind = ticker_to_industry.get(sym, "Các ngành khác")
                
                # Row 1: Industry name with custom background
                cell_ind = ws.cell(row=1, column=col_idx, value=ind)
                color_hex = industry_colors.get(ind, "366092")
                cell_ind.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                cell_ind.font = font_industry
                cell_ind.alignment = Alignment(horizontal="center", vertical="center")
                cell_ind.border = thin_border
                
                # Row 2: Ticker
                cell_sym = ws.cell(row=2, column=col_idx, value=sym)
                cell_sym.font = font_ticker
                cell_sym.alignment = Alignment(horizontal="center", vertical="center")
                cell_sym.border = thin_border
                
                # Row 3: Subheader
                sub_val = "CUR_MKT_CAP" if metric_type == "cap" else "LIQUIDITY"
                cell_sub = ws.cell(row=3, column=col_idx, value=sub_val)
                cell_sub.font = font_subheader
                cell_sub.alignment = Alignment(horizontal="center", vertical="center")
                cell_sub.border = thin_border
                
            # Write data rows
            for r_idx, dt in enumerate(dates, 4):
                try:
                    dt_obj = datetime.strptime(dt, "%Y-%m-%d")
                    formatted_dt = dt_obj.strftime("%d/%m/%Y")
                except:
                    formatted_dt = dt
                    
                cell_dt = ws.cell(row=r_idx, column=1, value=formatted_dt)
                cell_dt.font = font_dates
                cell_dt.alignment = Alignment(horizontal="center", vertical="center")
                cell_dt.border = thin_border
                
                day_prices = price_matrix.get(dt, {})
                day_volumes = volume_matrix.get(dt, {})
                
                for col_idx, sym in enumerate(ordered_tickers, 2):
                    cl = day_prices.get(sym)
                    vol = day_volumes.get(sym)
                    shares = get_outstanding_shares(sym, shares_map)

                    val = None
                    if cl is not None and cl > 0:
                        actual_price = normalize_price(cl)
                        if metric_type == "cap":
                            if shares:
                                val = (actual_price * shares) / 1_000_000_000
                        elif vol is not None:
                            val = (actual_price * int(vol)) / 1_000_000_000
                        
                    cell = ws.cell(row=r_idx, column=col_idx, value=round(val, 2) if val is not None else None)
                    cell.font = font_data
                    cell.number_format = '#,##0.0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = thin_border
                    
            # Merge Row 1 cells per industry block
            start_col = 2
            for col_idx in range(3, len(ordered_tickers) + 3):
                val_prev = ws.cell(row=1, column=col_idx - 1).value
                val_curr = ws.cell(row=1, column=col_idx).value
                if val_curr != val_prev or col_idx == len(ordered_tickers) + 2:
                    end_col = col_idx - 1
                    if end_col > start_col:
                        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    start_col = col_idx
                    
            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 18
            ws.column_dimensions[get_column_letter(1)].width = 16
            
            for col_idx in range(2, len(ordered_tickers) + 2):
                ws.column_dimensions[get_column_letter(col_idx)].width = 13
                
        ws_cap = wb.create_sheet(title="Vốn Hóa (Market Cap)")
        populate_worksheet(ws_cap, "cap")
        
        ws_liq = wb.create_sheet(title="Thanh Khoản (Liquidity)")
        populate_worksheet(ws_liq, "liq")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=market_grouped_range_{start_date}_to_{end_date}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/vol-cap/results")
def get_vol_cap_results(date: str = Query(...)):
    # Find nearest trading date if requested date has no data
    try:
        from datetime import datetime, timedelta
        dt_obj = datetime.strptime(date, '%Y-%m-%d')
        start_dt_str = (dt_obj - timedelta(days=15)).strftime('%Y-%m-%d')
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT time FROM historical_prices WHERE time >= ? AND time <= ? ORDER BY time DESC", (start_dt_str, date))
        dates = [row[0] for row in cursor.fetchall()]
        conn.close()
        if dates:
            date = dates[0]
    except Exception as e:
        print(f"Error finding nearest date for Vol/Cap: {e}")

    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
        shares_map = {row[0]: row[1] for row in cursor.fetchall()}
        
        cursor.execute("SELECT symbol, close, volume FROM historical_prices WHERE time = ?", (date,))
        rows = cursor.fetchall()
        conn.close()
        
        price_map = {r[0]: r[1] for r in rows}
        volume_map = {r[0]: r[2] for r in rows}
        
        results_list = []
        total_cap_billion = 0.0
        total_liq_vnd = 0.0
        
        for symbol in VN302:
            close_val = price_map.get(symbol)
            volume_val = volume_map.get(symbol)
            shares = get_outstanding_shares(symbol, shares_map)
            if close_val is None or close_val <= 0 or volume_val is None or not shares:
                continue
                
            actual_price = normalize_price(close_val)
            cap_vnd = actual_price * shares
            liq_vnd = actual_price * int(volume_val)
            
            total_cap_billion += cap_vnd / 1_000_000_000
            total_liq_vnd += liq_vnd
            
            vol_cap_pct = (liq_vnd / cap_vnd) * 100.0 if cap_vnd > 0 else 0.0
                
            industry = "Chưa phân loại"
            for ind, symbols in VN302_INDUSTRIES.items():
                if symbol in symbols:
                    industry = ind
                    break
                    
            results_list.append({
                "Ticker": symbol,
                "Industry": industry,
                "Close": actual_price,
                "Volume": int(volume_val),
                "OutstandingShares": shares,
                "LiquidityVND": liq_vnd,
                "MarketCapBillion": round(cap_vnd / 1_000_000_000, 2),
                "VolCapPct": round(vol_cap_pct, 4)
            })
            
        if results_list:
            results_list = sorted(results_list, key=lambda x: x["VolCapPct"], reverse=True)
            for rank, r in enumerate(results_list, 1):
                r["Rank"] = rank
                
            # Summaries
            avg_vol_cap = sum(x["VolCapPct"] for x in results_list) / len(results_list) if results_list else 0.0
            leader_ticker = results_list[0]["Ticker"] if results_list else "N/A"
            leader_pct = results_list[0]["VolCapPct"] if results_list else 0.0
            
            # Group sectors to populate sector dropdown
            sectors_summary = []
            ind_map = {}
            for r in results_list:
                ind = r["Industry"]
                if ind not in ind_map:
                    ind_map[ind] = 0
                ind_map[ind] += 1
            sectors_summary = [{"Industry": k, "Count": v} for k, v in ind_map.items()]
            
            return {
                "results": results_list,
                "date": date,
                "summary": {
                    "total_cap_billion": round(total_cap_billion, 2),
                    "total_liquidity_vnd": round(total_liq_vnd, 0),
                    "avg_vol_cap": round(avg_vol_cap, 4),
                    "leader_ticker": f"{leader_ticker} ({round(leader_pct, 2)}%)"
                },
                "sectors": sectors_summary
            }
        else:
            return {"results": [], "error": "Chưa có dữ liệu cho ngày này."}
            
    except Exception as e:
        return {"results": [], "error": str(e)}


@app.get("/api/export/vol-cap-history")
def export_vol_cap_history():
    try:
        from datetime import datetime, timedelta
        conn = get_db_conn()
        cursor = conn.cursor()
        
        # Find maximum date in database
        cursor.execute("SELECT MAX(time) FROM historical_prices")
        max_date_str = cursor.fetchone()[0]
        if not max_date_str:
            raise HTTPException(status_code=400, detail="Không tìm thấy dữ liệu trong database. Vui lòng bấm 'Đồng bộ Data 3 năm' trước.")
            
        max_dt = datetime.strptime(max_date_str, "%Y-%m-%d")
        start_dt = max_dt - timedelta(days=3*365)
        start_date = start_dt.strftime("%Y-%m-%d")
        
        # 1. Fetch dates in range
        cursor.execute("""
            SELECT DISTINCT time FROM historical_prices 
            WHERE time >= ? AND time <= ? 
            ORDER BY time DESC
        """, (start_date, max_date_str))
        dates = [row[0] for row in cursor.fetchall()]
        
        if not dates:
            raise HTTPException(status_code=400, detail="Không tìm thấy dữ liệu trong database.")
            
        # 2. Fetch outstanding shares
        cursor.execute("SELECT symbol, outstanding_shares FROM ticker_shares")
        shares_map = {row[0]: row[1] for row in cursor.fetchall()}
        
        # 3. Fetch historical close prices and volumes
        cursor.execute("""
            SELECT symbol, time, close, volume FROM historical_prices 
            WHERE time >= ? AND time <= ?
        """, (start_date, max_date_str))
        rows = cursor.fetchall()
        conn.close()
        
        price_matrix = {}
        volume_matrix = {}
        for sym, dt, cl, vol in rows:
            price_matrix.setdefault(dt, {})[sym] = cl
            volume_matrix.setdefault(dt, {})[sym] = vol
            
        # Group tickers by Industry horizontally
        ordered_tickers = []
        ticker_to_industry = {}
        for ind, syms in VN302_INDUSTRIES.items():
            for sym in syms:
                if sym in VN302:
                    ordered_tickers.append(sym)
                    ticker_to_industry[sym] = ind
        for sym in VN302:
            if sym not in ticker_to_industry:
                ordered_tickers.append(sym)
                ticker_to_industry[sym] = "Các ngành khác"
                
        # Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Tỷ lệ Vol-Cap"
        
        # Ensure grid lines are visible
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_industry = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_ticker = Font(name="Segoe UI", size=10, bold=True)
        font_subheader = Font(name="Segoe UI", size=9, italic=True, color="555555")
        font_data = Font(name="Segoe UI", size=10)
        font_dates = Font(name="Segoe UI", size=10, bold=True)
        
        industry_colors = {
            "Ngân hàng": "1F497D",  # Dark Blue
            "Bất động sản": "E46C0A", # Dark Orange
            "BĐS": "E46C0A",
            "Chứng khoán": "76933C", # Green
            "Dầu khí": "953734",     # Red-Brown
            "Thép": "1F4E79",        # Dark Blue (Steel blue)
            "Hóa chất": "7030A0",    # Deep Purple
            "Hàng không": "31859C",  # Light Teal
            "Công nghệ": "4F81BD",   # Blue
            "Năng lượng": "595959",  # Charcoal
            "Bán lẻ": "B9CDE5",
            "Xây dựng & Vật liệu": "D8E4BC",
            "Các ngành khác": "7F7F7F"
        }
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Set Dates Header
        ws.cell(row=3, column=1, value="Dates").font = font_dates
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=3, column=1).border = thin_border
        
        # Write headers
        for col_idx, sym in enumerate(ordered_tickers, 2):
            ind = ticker_to_industry.get(sym, "Các ngành khác")
            
            # Row 1: Industry Name with filled color
            cell_ind = ws.cell(row=1, column=col_idx, value=ind)
            color_hex = industry_colors.get(ind, "366092")
            cell_ind.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            cell_ind.font = font_industry
            cell_ind.alignment = Alignment(horizontal="center", vertical="center")
            cell_ind.border = thin_border
            
            # Row 2: Ticker
            cell_sym = ws.cell(row=2, column=col_idx, value=sym)
            cell_sym.font = font_ticker
            cell_sym.alignment = Alignment(horizontal="center", vertical="center")
            cell_sym.border = thin_border
            
            # Row 3: Subheader
            cell_sub = ws.cell(row=3, column=col_idx, value="Vol/Cap")
            cell_sub.font = font_subheader
            cell_sub.alignment = Alignment(horizontal="center", vertical="center")
            cell_sub.border = thin_border
            
        # Write Data
        for r_idx, dt in enumerate(dates, 4):
            try:
                dt_obj = datetime.strptime(dt, "%Y-%m-%d")
                formatted_dt = dt_obj.strftime("%d/%m/%Y")
            except:
                formatted_dt = dt
                
            cell_dt = ws.cell(row=r_idx, column=1, value=formatted_dt)
            cell_dt.font = font_dates
            cell_dt.alignment = Alignment(horizontal="center", vertical="center")
            cell_dt.border = thin_border
            
            day_prices = price_matrix.get(dt, {})
            day_volumes = volume_matrix.get(dt, {})
            
            for col_idx, sym in enumerate(ordered_tickers, 2):
                close_val = day_prices.get(sym)
                vol = day_volumes.get(sym)
                shares = get_outstanding_shares(sym, shares_map)

                ratio = None
                if close_val is not None and close_val > 0 and vol is not None and shares:
                    price_vnd = normalize_price(close_val)
                    liquidity_vnd = price_vnd * int(vol)
                    market_cap_vnd = price_vnd * shares
                    ratio = liquidity_vnd / market_cap_vnd if market_cap_vnd > 0 else None

                cell = ws.cell(row=r_idx, column=col_idx, value=ratio)
                cell.font = font_data
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
        # Merge Row 1 cells per industry block
        start_col = 2
        for col_idx in range(3, len(ordered_tickers) + 3):
            val_prev = ws.cell(row=1, column=col_idx - 1).value
            val_curr = ws.cell(row=1, column=col_idx).value
            if val_curr != val_prev or col_idx == len(ordered_tickers) + 2:
                end_col = col_idx - 1
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                start_col = col_idx
                
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18
        ws.column_dimensions[get_column_letter(1)].width = 16
        
        for col_idx in range(2, len(ordered_tickers) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=vol_cap_3years_history.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.on_event("startup")
def on_startup():
    init_db_liquidity()
    init_db_shares()
    # Bypass crawlers in serverless environments to prevent request timeouts
    if os.getenv("VERCEL") != "1":
        start_background_crawler()
        start_shares_crawler()

# Mount Static Files (Must be registered last to avoid route conflicts)
app.mount("/", StaticFiles(directory="static"), name="static")

if __name__ == "__main__":
    import uvicorn
    # Launch uvicorn server on port 8000
    uvicorn.run("web_server:app", host="127.0.0.1", port=8000, reload=True)
